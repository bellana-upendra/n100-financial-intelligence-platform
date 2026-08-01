"""One-command builder for Sprint 4: Streamlit Dashboard + Valuation.

Run this file from the project root:
    python build_sprint4_all.py

It creates/updates:
- src/dashboard/app.py
- src/dashboard/utils/db.py
- src/dashboard/utils/formatting.py
- src/dashboard/pages/01_home.py ... 08_reports.py
- src/analytics/valuation.py
- run_sprint4.bat
- README.md Sprint 4 section
- output/sprint4_retrospective.md

Existing target files are backed up before replacement.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
import py_compile
import shutil
import sys

ROOT = Path.cwd().resolve()
STAMP = datetime.now().strftime("%Y%m%d_%H%M%S")


def require_project_root() -> None:
    required = [ROOT / "src", ROOT / "src" / "config.py"]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise SystemExit(
            "Run this script from the N100 project root. Missing:\n- "
            + "\n- ".join(missing)
        )


def backup_and_write(relative_path: str, content: str) -> None:
    path = ROOT / relative_path
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists() and path.read_text(encoding="utf-8", errors="ignore") != content:
        backup = path.with_suffix(path.suffix + f".bak_{STAMP}")
        shutil.copy2(path, backup)
        print(f"Backup: {backup.relative_to(ROOT)}")
    path.write_text(content, encoding="utf-8")
    print(f"Written: {relative_path}")


def touch(relative_path: str) -> None:
    path = ROOT / relative_path
    path.parent.mkdir(parents=True, exist_ok=True)
    path.touch(exist_ok=True)


APP_PY = r'''"""Main entry point for the Nifty 100 Streamlit dashboard."""

from pathlib import Path
import sys

import streamlit as st

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

st.set_page_config(
    page_title="Nifty 100 Analytics",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

PAGE_DIR = Path(__file__).resolve().parent / "pages"

if hasattr(st, "navigation"):
    pages = [
        st.Page(PAGE_DIR / "01_home.py", title="Home / Overview", icon="🏠", default=True),
        st.Page(PAGE_DIR / "02_profile.py", title="Company Profile", icon="🏢"),
        st.Page(PAGE_DIR / "03_screener.py", title="Financial Screener", icon="🔎"),
        st.Page(PAGE_DIR / "04_peers.py", title="Peer Comparison", icon="⚖️"),
        st.Page(PAGE_DIR / "05_trends.py", title="Trend Analysis", icon="📈"),
        st.Page(PAGE_DIR / "06_sectors.py", title="Sector Analysis", icon="🧩"),
        st.Page(PAGE_DIR / "07_capital.py", title="Capital Allocation", icon="🗺️"),
        st.Page(PAGE_DIR / "08_reports.py", title="Annual Reports", icon="📄"),
    ]
    navigation = st.navigation(pages)
    navigation.run()
else:
    st.title("Nifty 100 Financial Intelligence Platform")
    st.info(
        "Use the sidebar to open the eight dashboard screens. "
        "Upgrade Streamlit for the enhanced navigation menu."
    )
'''

DB_PY = r'''"""Cached SQLite data-access layer for all Streamlit screens."""

from __future__ import annotations

from pathlib import Path
import sqlite3

import pandas as pd
import streamlit as st

from src.config import get_settings

SETTINGS = get_settings()
DATABASE_PATH = Path(SETTINGS.database_path)


def _read_query(query: str, params: tuple = ()) -> pd.DataFrame:
    """Execute a parameterised query using a fresh SQLite connection."""
    if not DATABASE_PATH.exists():
        raise FileNotFoundError(f"Database not found: {DATABASE_PATH}")
    with sqlite3.connect(DATABASE_PATH) as connection:
        connection.execute("PRAGMA foreign_keys = ON")
        return pd.read_sql_query(query, connection, params=params)


@st.cache_data(ttl=600)
def get_companies() -> pd.DataFrame:
    return _read_query(
        """
        SELECT
            c.company_id AS company_id,
            c.company_name,
            c.company_logo,
            c.about_company,
            c.website,
            c.nse_profile,
            c.bse_profile,
            c.face_value,
            c.book_value,
            c.roce_percentage AS source_roce_percentage,
            c.roe_percentage AS source_roe_percentage,
            c.broad_sector,
            c.sub_sector,
            c.index_weight_pct,
            c.market_cap_category
        FROM companies c
        ORDER BY c.company_name
        """
    )


@st.cache_data(ttl=600)
def get_company(ticker: str) -> pd.DataFrame:
    return _read_query(
        """
        SELECT
            c.company_id AS company_id,
            c.company_name,
            c.company_logo,
            c.about_company,
            c.website,
            c.nse_profile,
            c.bse_profile,
            c.face_value,
            c.book_value,
            c.broad_sector,
            c.sub_sector,
            c.market_cap_category
        FROM companies c
        WHERE c.company_id = ?
        """,
        (ticker.strip().upper(),),
    )


@st.cache_data(ttl=600)
def get_ratios(ticker: str, year: int | None = None) -> pd.DataFrame:
    ticker = ticker.strip().upper()
    if year is None:
        return _read_query(
            """
            SELECT *
            FROM financial_ratios
            WHERE company_id = ?
            ORDER BY CAST(year AS INTEGER)
            """,
            (ticker,),
        )
    return _read_query(
        """
        SELECT *
        FROM financial_ratios
        WHERE company_id = ? AND CAST(year AS INTEGER) = ?
        ORDER BY id
        """,
        (ticker, int(year)),
    )


@st.cache_data(ttl=600)
def get_pl(ticker: str) -> pd.DataFrame:
    return _read_query(
        """
        SELECT * FROM profitandloss
        WHERE company_id = ?
        ORDER BY CAST(year AS INTEGER)
        """,
        (ticker.strip().upper(),),
    )


@st.cache_data(ttl=600)
def get_bs(ticker: str) -> pd.DataFrame:
    return _read_query(
        """
        SELECT * FROM balancesheet
        WHERE company_id = ?
        ORDER BY CAST(year AS INTEGER)
        """,
        (ticker.strip().upper(),),
    )


@st.cache_data(ttl=600)
def get_cf(ticker: str) -> pd.DataFrame:
    return _read_query(
        """
        SELECT * FROM cashflow
        WHERE company_id = ?
        ORDER BY CAST(year AS INTEGER)
        """,
        (ticker.strip().upper(),),
    )


@st.cache_data(ttl=600)
def get_sectors() -> pd.DataFrame:
    return _read_query(
        """
        SELECT
	    company_id,
	    broad_sector,
	    sub_sector,
	    index_weight_pct,
	    market_cap_category
	FROM companies
        ORDER BY broad_sector, sub_sector, company_id
        """
    )


@st.cache_data(ttl=600)
def get_peer_group_names() -> list[str]:
    data = _read_query(
        """
        SELECT DISTINCT peer_group_name
        FROM peer_groups
        ORDER BY peer_group_name
        """
    )
    return data["peer_group_name"].dropna().tolist()


@st.cache_data(ttl=600)
def get_peers(group_name: str) -> pd.DataFrame:
    return _read_query(
        """
        SELECT pg.*, c.company_name
        FROM peer_groups pg
        LEFT JOIN companies c ON pg.company_id = c.company_id
        WHERE pg.peer_group_name = ?
        ORDER BY pg.is_benchmark DESC, c.company_name
        """,
        (group_name,),
    )


@st.cache_data(ttl=600)
def get_peer_years(group_name: str) -> list[int]:
    data = _read_query(
        """
        SELECT DISTINCT CAST(r.year AS INTEGER) AS year
        FROM financial_ratios r
        JOIN peer_groups pg ON r.company_id = pg.company_id
        WHERE pg.peer_group_name = ?
        ORDER BY year DESC
        """,
        (group_name,),
    )
    return [int(v) for v in data["year"].dropna().tolist()]


@st.cache_data(ttl=600)
def get_peer_group_data(group_name: str, year: int | None = None) -> pd.DataFrame:
    if year is None:
        return _read_query(
            """
            WITH ranked AS (
                SELECT r.*,
                       ROW_NUMBER() OVER (
                           PARTITION BY r.company_id
                           ORDER BY CAST(r.year AS INTEGER) DESC, r.id DESC
                       ) AS rn
                FROM financial_ratios r
            )
            SELECT
                pg.peer_group_name,
                pg.company_id,
                pg.is_benchmark,
                c.company_name,
                CAST(ranked.year AS INTEGER) AS financial_year,
                ranked.return_on_equity_pct,
                ranked.return_on_capital_employed_pct,
                ranked.net_profit_margin_pct,
                ranked.debt_to_equity,
                ranked.free_cash_flow_cr,
                ranked.pat_cagr_5yr,
                ranked.revenue_cagr_5yr,
                ranked.eps_cagr_5yr,
                ranked.interest_coverage,
                ranked.asset_turnover,
                ranked.composite_quality_score
            FROM peer_groups pg
            JOIN companies c ON pg.company_id = c.company_id
            LEFT JOIN ranked ON pg.company_id = ranked.company_id AND ranked.rn = 1
            WHERE pg.peer_group_name = ?
            ORDER BY pg.is_benchmark DESC, c.company_name
            """,
            (group_name,),
        )
    return _read_query(
        """
        WITH ratios AS (
            SELECT *
            FROM financial_ratios
            WHERE CAST(year AS INTEGER) = ?
        )
        SELECT
            pg.peer_group_name,
            pg.company_id,
            pg.is_benchmark,
            c.company_name,
            CAST(ratios.year AS INTEGER) AS financial_year,
            ratios.return_on_equity_pct,
            ratios.return_on_capital_employed_pct,
            ratios.net_profit_margin_pct,
            ratios.debt_to_equity,
            ratios.free_cash_flow_cr,
            ratios.pat_cagr_5yr,
            ratios.revenue_cagr_5yr,
            ratios.eps_cagr_5yr,
            ratios.interest_coverage,
            ratios.asset_turnover,
            ratios.composite_quality_score
        FROM peer_groups pg
        JOIN companies c ON pg.company_id = c.company_id
        LEFT JOIN ratios ON pg.company_id = ratios.company_id
        WHERE pg.peer_group_name = ?
        ORDER BY pg.is_benchmark DESC, c.company_name
        """,
        (int(year), group_name),
    )


@st.cache_data(ttl=600)
def get_valuation(ticker: str) -> pd.DataFrame:
    return _read_query(
        """
        SELECT
            m.*,
            c.company_name,
            c.broad_sector,
            c.sub_sector
        FROM market_cap m
        LEFT JOIN companies c ON m.company_id = c.company_id
        WHERE m.company_id = ?
        ORDER BY CAST(m.year AS INTEGER)
        """,
        (ticker.strip().upper(),),
    )


@st.cache_data(ttl=600)
def get_home_data(year: int) -> pd.DataFrame:
    return _read_query(
        """
        WITH ratios AS (
            SELECT
                company_id,
                CAST(year AS INTEGER) AS year,
                AVG(return_on_equity_pct) AS return_on_equity_pct,
                AVG(debt_to_equity) AS debt_to_equity,
                AVG(revenue_cagr_5yr) AS revenue_cagr_5yr,
                AVG(composite_quality_score) AS composite_quality_score
            FROM financial_ratios
            WHERE CAST(year AS INTEGER) = ?
            GROUP BY company_id, CAST(year AS INTEGER)
        ), market AS (
            SELECT * FROM market_cap WHERE CAST(year AS INTEGER) = ?
        )
        SELECT
            c.company_id AS company_id,
            c.company_name,
            c.broad_sector,
            c.sub_sector,
            ratios.return_on_equity_pct,
            ratios.debt_to_equity,
            ratios.revenue_cagr_5yr,
            ratios.composite_quality_score,
            market.pe_ratio,
            market.pb_ratio,
            market.dividend_yield_pct,
            market.market_cap_crore
        FROM companies c
        LEFT JOIN ratios ON c.company_id = ratios.company_id
        LEFT JOIN market ON c.company_id = market.company_id
        ORDER BY c.company_name
        """,
        (int(year), int(year)),
    )


@st.cache_data(ttl=600)
def get_screener_data() -> pd.DataFrame:
    return _read_query(
        """
        WITH latest_ratio AS (
            SELECT * FROM (
                SELECT r.*,
                       ROW_NUMBER() OVER (
                           PARTITION BY r.company_id
                           ORDER BY CAST(r.year AS INTEGER) DESC, r.id DESC
                       ) AS rn
                FROM financial_ratios r
            ) WHERE rn = 1
        ), latest_market AS (
            SELECT * FROM (
                SELECT m.*,
                       ROW_NUMBER() OVER (
                           PARTITION BY m.company_id
                           ORDER BY CAST(m.year AS INTEGER) DESC, m.id DESC
                       ) AS rn
                FROM market_cap m
            ) WHERE rn = 1
        )
        SELECT
            c.company_id AS company_id,
            c.company_name,
            c.broad_sector,
            c.sub_sector,
            latest_ratio.year,
            latest_ratio.return_on_equity_pct,
            latest_ratio.return_on_capital_employed_pct,
            latest_ratio.net_profit_margin_pct,
            latest_ratio.debt_to_equity,
            latest_ratio.free_cash_flow_cr,
            latest_ratio.revenue_cagr_5yr,
            latest_ratio.pat_cagr_5yr,
            latest_ratio.operating_profit_margin_pct,
            latest_ratio.interest_coverage,
            latest_ratio.icr_label,
            latest_ratio.composite_quality_score,
            latest_ratio.dividend_payout_ratio_pct,
            latest_ratio.capital_allocation_pattern,
            latest_market.pe_ratio,
            latest_market.pb_ratio,
            latest_market.ev_ebitda,
            latest_market.dividend_yield_pct,
            latest_market.market_cap_crore
        FROM companies c
        LEFT JOIN latest_ratio ON c.company_id = latest_ratio.company_id
        LEFT JOIN latest_market ON c.company_id = latest_market.company_id
        ORDER BY c.company_name
        """
    )


@st.cache_data(ttl=600)
def get_pros_cons(ticker: str) -> pd.DataFrame:
    return _read_query(
        """
        SELECT pros, cons
        FROM prosandcons
        WHERE company_id = ?
        ORDER BY id
        """,
        (ticker.strip().upper(),),
    )


@st.cache_data(ttl=600)
def get_documents(ticker: str) -> pd.DataFrame:
    return _read_query(
        """
        SELECT id, company_id, Year AS report_year, Annual_Report AS report_url
        FROM documents
        WHERE company_id = ?
        ORDER BY Year DESC
        """,
        (ticker.strip().upper(),),
    )


@st.cache_data(ttl=600)
def get_sector_names() -> list[str]:
    data = _read_query(
        """
        SELECT DISTINCT broad_sector
        FROM companies
        WHERE broad_sector IS NOT NULL
          AND TRIM(broad_sector) <> ''
        ORDER BY broad_sector
        """
    )
    return data["broad_sector"].tolist()


@st.cache_data(ttl=600)
def get_sector_data(sector: str, year: int) -> pd.DataFrame:
    return _read_query(
        """
        WITH ratios AS (
            SELECT
                company_id,
                CAST(year AS INTEGER) AS year,
                AVG(return_on_equity_pct) AS return_on_equity_pct,
                AVG(return_on_capital_employed_pct) AS return_on_capital_employed_pct,
                AVG(net_profit_margin_pct) AS net_profit_margin_pct,
                AVG(debt_to_equity) AS debt_to_equity,
                AVG(revenue_cagr_5yr) AS revenue_cagr_5yr,
                AVG(pat_cagr_5yr) AS pat_cagr_5yr,
                AVG(free_cash_flow_cr) AS free_cash_flow_cr
            FROM financial_ratios
            WHERE CAST(year AS INTEGER) = ?
            GROUP BY company_id, CAST(year AS INTEGER)
        ), pl AS (
            SELECT company_id, CAST(year AS INTEGER) AS year, AVG(sales) AS sales
            FROM profitandloss
            WHERE CAST(year AS INTEGER) = ?
            GROUP BY company_id, CAST(year AS INTEGER)
        ), market AS (
            SELECT * FROM market_cap WHERE CAST(year AS INTEGER) = ?
        )
        SELECT
            c.company_id AS company_id,
            c.company_name,
            c.broad_sector,
            c.sub_sector,
            pl.sales,
            ratios.return_on_equity_pct,
            ratios.return_on_capital_employed_pct,
            ratios.net_profit_margin_pct,
            ratios.debt_to_equity,
            ratios.revenue_cagr_5yr,
            ratios.pat_cagr_5yr,
            ratios.free_cash_flow_cr,
            market.market_cap_crore,
            market.pe_ratio,
            market.pb_ratio,
            market.ev_ebitda,
            market.dividend_yield_pct
        FROM companies c
        LEFT JOIN ratios ON c.company_id = ratios.company_id
        LEFT JOIN pl ON c.company_id = pl.company_id
        LEFT JOIN market ON c.company_id = market.company_id
        WHERE c.broad_sector = ?
        ORDER BY c.company_name
        """,
        (int(year), int(year), int(year), sector),
    )


@st.cache_data(ttl=600)
def get_capital_allocation_data() -> pd.DataFrame:
    return _read_query(
        """
        WITH latest_ratio AS (
            SELECT * FROM (
                SELECT r.*,
                       ROW_NUMBER() OVER (
                           PARTITION BY r.company_id
                           ORDER BY CAST(r.year AS INTEGER) DESC, r.id DESC
                       ) AS rn
                FROM financial_ratios r
            ) WHERE rn = 1
        ), latest_market AS (
            SELECT * FROM (
                SELECT m.*,
                       ROW_NUMBER() OVER (
                           PARTITION BY m.company_id
                           ORDER BY CAST(m.year AS INTEGER) DESC, m.id DESC
                       ) AS rn
                FROM market_cap m
            ) WHERE rn = 1
        )
        SELECT
            c.company_id AS company_id,
            c.company_name,
            c.broad_sector,
            c.sub_sector,
            latest_ratio.year,
            latest_ratio.cfo_sign,
            latest_ratio.cfi_sign,
            latest_ratio.cff_sign,
            latest_ratio.capital_allocation_pattern,
            latest_ratio.free_cash_flow_cr,
            latest_ratio.cfo_quality_label,
            latest_market.market_cap_crore
        FROM companies c
        LEFT JOIN latest_ratio ON c.company_id = latest_ratio.company_id
        LEFT JOIN latest_market ON c.company_id = latest_market.company_id
        ORDER BY c.company_name
        """
    )
'''

FORMATTING_PY = r'''"""Display and data-cleaning helpers for dashboard pages."""

from __future__ import annotations

import math
import pandas as pd


def format_metric(value, suffix: str = "", decimals: int = 2) -> str:
    if value is None:
        return "N/A"
    try:
        if pd.isna(value) or not math.isfinite(float(value)):
            return "N/A"
        return f"{float(value):,.{decimals}f}{suffix}"
    except (TypeError, ValueError):
        return str(value)


def latest_value(data: pd.DataFrame, column: str):
    if data.empty or column not in data.columns:
        return None
    values = pd.to_numeric(data[column], errors="coerce").dropna()
    return None if values.empty else values.iloc[-1]


def clean_year(data: pd.DataFrame) -> pd.DataFrame:
    result = data.copy()
    if "year" in result.columns:
        result["year"] = pd.to_numeric(result["year"], errors="coerce")
        result = result.dropna(subset=["year"])
        result["year"] = result["year"].astype(int)
    return result
'''

PAGE_COMMON_HEADER = r'''from pathlib import Path
import sys

PROJECT_ROOT = Path(__file__).resolve().parents[3]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))
'''

HOME_PY = PAGE_COMMON_HEADER + r'''
import pandas as pd
import plotly.express as px
import streamlit as st

from src.dashboard.utils.db import get_companies, get_home_data
from src.dashboard.utils.formatting import format_metric

st.title("Home / Overview")
st.caption("Nifty 100 market summary, sector distribution and quality leaders")

year = st.sidebar.selectbox("Financial year", list(range(2024, 2018, -1)), key="home_year")
data = get_home_data(year)

if data.empty:
    st.warning("No dashboard data is available for the selected year.")
    st.stop()

avg_roe = pd.to_numeric(data["return_on_equity_pct"], errors="coerce").mean()
median_pe = pd.to_numeric(data["pe_ratio"], errors="coerce").median()
median_de = pd.to_numeric(data["debt_to_equity"], errors="coerce").median()
total_companies = data["company_id"].nunique()
median_cagr = pd.to_numeric(data["revenue_cagr_5yr"], errors="coerce").median()
debt_free = int((pd.to_numeric(data["debt_to_equity"], errors="coerce") == 0).sum())

row1 = st.columns(3)
row2 = st.columns(3)
row1[0].metric("Average ROE", format_metric(avg_roe, "%"))
row1[1].metric("Median P/E", format_metric(median_pe, "x"))
row1[2].metric("Median D/E", format_metric(median_de, "x"))
row2[0].metric("Total Companies", f"{total_companies}")
row2[1].metric("Median Revenue CAGR 5yr", format_metric(median_cagr, "%"))
row2[2].metric("Debt-Free Companies", f"{debt_free}")

left, right = st.columns([1.15, 1])
with left:
    companies = get_companies()
    sector_counts = (
        companies.assign(broad_sector=companies["broad_sector"].fillna("Unclassified"))
        .groupby("broad_sector", as_index=False)
        .agg(company_count=("company_id", "nunique"))
    )
    fig = px.pie(
        sector_counts,
        names="broad_sector",
        values="company_count",
        hole=0.55,
        title="Sector Breakdown",
    )
    fig.update_layout(margin=dict(l=10, r=10, t=55, b=10))
    st.plotly_chart(fig, use_container_width=True)

with right:
    st.subheader("Top 5 by Composite Quality Score")
    top = data.copy()
    top["composite_quality_score"] = pd.to_numeric(
        top["composite_quality_score"], errors="coerce"
    )
    top = top.dropna(subset=["composite_quality_score"]).nlargest(
        5, "composite_quality_score"
    )
    display = top[
        ["company_id", "company_name", "broad_sector", "composite_quality_score"]
    ].rename(
        columns={
            "company_id": "Ticker",
            "company_name": "Company",
            "broad_sector": "Sector",
            "composite_quality_score": "Quality Score",
        }
    )
    st.dataframe(display, use_container_width=True, hide_index=True)

missing_ratio = int(data["return_on_equity_pct"].isna().sum())
if missing_ratio:
    st.info(
        f"{missing_ratio} companies do not have a ratio record for {year}; "
        "available companies are used in calculated KPIs."
    )
'''

PROFILE_PY = PAGE_COMMON_HEADER + r'''
import time

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

from src.dashboard.utils.db import (
    get_cf,
    get_companies,
    get_company,
    get_pl,
    get_pros_cons,
    get_ratios,
)
from src.dashboard.utils.formatting import clean_year, format_metric, latest_value

started = time.perf_counter()
st.title("Company Profile")
st.caption("Company overview, financial KPIs, ten-year performance and qualitative insights")

companies = get_companies()
labels = {
    row.company_id: f"{row.company_name} — {row.company_id}"
    for row in companies.itertuples()
}
tickers = companies["company_id"].tolist()
selected = st.selectbox(
    "Search company name or ticker",
    options=tickers,
    format_func=lambda ticker: labels.get(ticker, ticker),
    key="profile_company",
)

company_df = get_company(selected)
if company_df.empty:
    st.warning("Ticker not found — please try another.")
    st.stop()
company = company_df.iloc[0]

ratios = clean_year(get_ratios(selected))
pl = clean_year(get_pl(selected))
cf = clean_year(get_cf(selected))
pros_cons = get_pros_cons(selected)

header_left, header_right = st.columns([1, 5])
with header_left:
    logo = company.get("company_logo")
    if isinstance(logo, str) and logo.strip():
        st.image(logo, width=110)
with header_right:
    st.subheader(str(company.get("company_name", selected)))
    st.markdown(
        f"**Ticker:** {selected}  |  **Sector:** {company.get('broad_sector') or 'N/A'}  "
        f"|  **Sub-sector:** {company.get('sub_sector') or 'N/A'}"
    )
    description = company.get("about_company")
    st.write(description if isinstance(description, str) and description.strip() else "Description unavailable.")
    website = company.get("website")
    if isinstance(website, str) and website.strip():
        st.link_button("Company Website", website)

latest_ratios = ratios.sort_values("year").tail(1)
latest = latest_ratios.iloc[0] if not latest_ratios.empty else pd.Series(dtype=object)

kpis = [
    ("ROE", latest.get("return_on_equity_pct"), "%"),
    ("ROCE", latest.get("return_on_capital_employed_pct"), "%"),
    ("Net Profit Margin", latest.get("net_profit_margin_pct"), "%"),
    ("Debt-to-Equity", latest.get("debt_to_equity"), "x"),
    ("Revenue CAGR 5yr", latest.get("revenue_cagr_5yr"), "%"),
    ("Free Cash Flow", latest.get("free_cash_flow_cr"), " Cr"),
]
cols = st.columns(6)
for col, (label, value, suffix) in zip(cols, kpis):
    col.metric(label, format_metric(value, suffix))

chart_pl = pl.sort_values("year").drop_duplicates("year", keep="last").tail(10)
if chart_pl.empty:
    st.info("Revenue and profit data are unavailable for this company.")
else:
    available_years = chart_pl["year"].nunique()
    if available_years < 10:
        st.info(f"Only {available_years} years of P&L data are available for this company.")
    long_pl = chart_pl.melt(
        id_vars="year",
        value_vars=["sales", "net_profit"],
        var_name="Metric",
        value_name="₹ Crore",
    )
    long_pl["Metric"] = long_pl["Metric"].map(
        {"sales": "Revenue", "net_profit": "Net Profit"}
    )
    fig = px.bar(
        long_pl,
        x="year",
        y="₹ Crore",
        color="Metric",
        barmode="group",
        title="Revenue and Net Profit — Latest 10 Years",
    )
    st.plotly_chart(fig, use_container_width=True)

trend = ratios.sort_values("year").drop_duplicates("year", keep="last").tail(10)
if not trend.empty:
    fig = go.Figure()
    fig.add_trace(
        go.Scatter(
            x=trend["year"],
            y=trend["return_on_equity_pct"],
            name="ROE",
            mode="lines+markers",
        )
    )
    fig.add_trace(
        go.Scatter(
            x=trend["year"],
            y=trend["return_on_capital_employed_pct"],
            name="ROCE",
            mode="lines+markers",
            yaxis="y2",
        )
    )
    fig.update_layout(
        title="ROE and ROCE Trend",
        xaxis_title="Financial Year",
        yaxis=dict(title="ROE (%)"),
        yaxis2=dict(title="ROCE (%)", overlaying="y", side="right"),
        legend=dict(orientation="h"),
    )
    st.plotly_chart(fig, use_container_width=True)
else:
    st.info("ROE and ROCE history are unavailable.")

pros = [str(v).strip() for v in pros_cons.get("pros", pd.Series(dtype=str)).dropna() if str(v).strip()]
cons = [str(v).strip() for v in pros_cons.get("cons", pd.Series(dtype=str)).dropna() if str(v).strip()]

roe = latest_value(latest_ratios, "return_on_equity_pct")
de = latest_value(latest_ratios, "debt_to_equity")
fcf = latest_value(latest_ratios, "free_cash_flow_cr")
rev_cagr = latest_value(latest_ratios, "revenue_cagr_5yr")
sector = company.get("broad_sector")

if not pros:
    if roe is not None and roe >= 15:
        pros.append("Strong latest-year return on equity")
    if fcf is not None and fcf > 0:
        pros.append("Positive free cash flow in the latest available year")
    if de is not None and de <= 1:
        pros.append("Conservative debt-to-equity level")
    if rev_cagr is not None and rev_cagr >= 10:
        pros.append("Healthy five-year revenue growth")
if not cons:
    if de is not None and de > 2 and sector != "Financials":
        cons.append("Elevated debt-to-equity level")
    if fcf is not None and fcf < 0:
        cons.append("Negative free cash flow in the latest available year")
    if rev_cagr is not None and rev_cagr < 5:
        cons.append("Five-year revenue growth is below 5%")
    if roe is not None and roe < 10:
        cons.append("Latest return on equity is below 10%")

st.subheader("Pros and Cons")
pro_col, con_col = st.columns(2)
with pro_col:
    st.markdown("#### Pros")
    if pros:
        for item in pros:
            st.success(f"✅ {item}")
    else:
        st.info("No positive observations are available.")
with con_col:
    st.markdown("#### Cons")
    if cons:
        for item in cons:
            st.error(f"❌ {item}")
    else:
        st.info("No risk observations are available.")

st.caption(f"Profile loaded in {time.perf_counter() - started:.2f} seconds")
'''

SCREENER_PY = PAGE_COMMON_HEADER + r'''
import numpy as np
import pandas as pd
import streamlit as st

from src.dashboard.utils.db import get_screener_data

st.title("Financial Screener")
st.caption("Live ten-metric screening with six ready-to-use presets and CSV export")

data = get_screener_data().copy()

PRESETS = {
    "Quality": dict(roe_min=15.0, de_max=1.0, fcf_min=0.0, revenue_cagr_min=10.0,
                    pat_cagr_min=-100.0, opm_min=0.0, pe_max=200.0, pb_max=30.0,
                    dividend_yield_min=0.0, icr_min=0.0),
    "Value": dict(roe_min=-100.0, de_max=2.0, fcf_min=-200000.0, revenue_cagr_min=-100.0,
                  pat_cagr_min=-100.0, opm_min=-100.0, pe_max=20.0, pb_max=3.0,
                  dividend_yield_min=1.0, icr_min=0.0),
    "Growth": dict(roe_min=-100.0, de_max=2.0, fcf_min=-200000.0, revenue_cagr_min=15.0,
                   pat_cagr_min=20.0, opm_min=-100.0, pe_max=200.0, pb_max=30.0,
                   dividend_yield_min=0.0, icr_min=0.0),
    "Dividend": dict(roe_min=-100.0, de_max=20.0, fcf_min=0.0, revenue_cagr_min=-100.0,
                     pat_cagr_min=-100.0, opm_min=-100.0, pe_max=200.0, pb_max=30.0,
                     dividend_yield_min=2.0, icr_min=0.0),
    "Debt-Free": dict(roe_min=12.0, de_max=0.0, fcf_min=-200000.0, revenue_cagr_min=-100.0,
                      pat_cagr_min=-100.0, opm_min=-100.0, pe_max=200.0, pb_max=30.0,
                      dividend_yield_min=0.0, icr_min=0.0),
    "Turnaround": dict(roe_min=-100.0, de_max=20.0, fcf_min=0.0, revenue_cagr_min=0.0,
                       pat_cagr_min=-100.0, opm_min=-100.0, pe_max=200.0, pb_max=30.0,
                       dividend_yield_min=0.0, icr_min=0.0),
}

DEFAULTS = dict(roe_min=-100.0, de_max=20.0, fcf_min=-200000.0, revenue_cagr_min=-100.0,
                pat_cagr_min=-100.0, opm_min=-100.0, pe_max=200.0, pb_max=30.0,
                dividend_yield_min=0.0, icr_min=0.0)
for key, value in DEFAULTS.items():
    st.session_state.setdefault(key, value)
st.session_state.setdefault("active_preset", "Custom")

st.sidebar.markdown("### Presets")
button_cols = st.sidebar.columns(2)
for index, (preset_name, values) in enumerate(PRESETS.items()):
    if button_cols[index % 2].button(preset_name, use_container_width=True, key=f"preset_{preset_name}"):
        for key, value in values.items():
            st.session_state[key] = value
        st.session_state["active_preset"] = preset_name
        st.rerun()
if st.sidebar.button("Reset", use_container_width=True):
    for key, value in DEFAULTS.items():
        st.session_state[key] = value
    st.session_state["active_preset"] = "Custom"
    st.rerun()

st.sidebar.caption(f"Active preset: {st.session_state['active_preset']}")
st.sidebar.markdown("### Metric Filters")
roe_min = st.sidebar.slider("ROE minimum (%)", -100.0, 200.0, key="roe_min")
de_max = st.sidebar.slider("D/E maximum", 0.0, 20.0, key="de_max")
fcf_min = st.sidebar.slider("FCF minimum (₹ Cr)", -200000.0, 200000.0, step=1000.0, key="fcf_min")
revenue_cagr_min = st.sidebar.slider("Revenue CAGR 5yr minimum (%)", -100.0, 200.0, key="revenue_cagr_min")
pat_cagr_min = st.sidebar.slider("PAT CAGR 5yr minimum (%)", -100.0, 200.0, key="pat_cagr_min")
opm_min = st.sidebar.slider("OPM minimum (%)", -100.0, 100.0, key="opm_min")
pe_max = st.sidebar.slider("P/E maximum", 0.0, 200.0, key="pe_max")
pb_max = st.sidebar.slider("P/B maximum", 0.0, 30.0, key="pb_max")
dividend_yield_min = st.sidebar.slider("Dividend Yield minimum (%)", 0.0, 10.0, key="dividend_yield_min")
icr_min = st.sidebar.slider("ICR minimum", 0.0, 100.0, key="icr_min")

numeric_columns = [
    "return_on_equity_pct", "debt_to_equity", "free_cash_flow_cr",
    "revenue_cagr_5yr", "pat_cagr_5yr", "operating_profit_margin_pct",
    "pe_ratio", "pb_ratio", "dividend_yield_pct", "interest_coverage",
]
for column in numeric_columns:
    data[column] = pd.to_numeric(data[column], errors="coerce")

mask = pd.Series(True, index=data.index)
mask &= data["return_on_equity_pct"].notna() & (data["return_on_equity_pct"] >= roe_min)

if st.session_state["active_preset"] == "Debt-Free":
    mask &= data["debt_to_equity"].notna() & np.isclose(data["debt_to_equity"], 0.0)
else:
    financials = data["broad_sector"].eq("Financials")
    mask &= financials | (data["debt_to_equity"].notna() & (data["debt_to_equity"] <= de_max))

mask &= data["free_cash_flow_cr"].notna() & (data["free_cash_flow_cr"] >= fcf_min)
mask &= data["revenue_cagr_5yr"].notna() & (data["revenue_cagr_5yr"] >= revenue_cagr_min)
mask &= data["pat_cagr_5yr"].notna() & (data["pat_cagr_5yr"] >= pat_cagr_min)
mask &= data["operating_profit_margin_pct"].notna() & (data["operating_profit_margin_pct"] >= opm_min)
mask &= data["pe_ratio"].notna() & (data["pe_ratio"] <= pe_max)
mask &= data["pb_ratio"].notna() & (data["pb_ratio"] <= pb_max)
mask &= data["dividend_yield_pct"].notna() & (data["dividend_yield_pct"] >= dividend_yield_min)

debt_free_icr = data["debt_to_equity"].fillna(np.inf).eq(0) | data["icr_label"].eq("Debt Free")
mask &= debt_free_icr | (data["interest_coverage"].notna() & (data["interest_coverage"] >= icr_min))

results = data.loc[mask].copy()
results["composite_quality_score"] = pd.to_numeric(results["composite_quality_score"], errors="coerce")
results = results.sort_values("composite_quality_score", ascending=False, na_position="last")

st.subheader(f"{len(results)} companies match your filters")
visible = results[
    [
        "company_id", "company_name", "broad_sector", "composite_quality_score",
        "return_on_equity_pct", "debt_to_equity", "free_cash_flow_cr",
        "revenue_cagr_5yr", "pat_cagr_5yr", "operating_profit_margin_pct",
        "pe_ratio", "pb_ratio", "dividend_yield_pct", "interest_coverage",
    ]
].rename(
    columns={
        "company_id": "Ticker", "company_name": "Company", "broad_sector": "Sector",
        "composite_quality_score": "Composite Score", "return_on_equity_pct": "ROE %",
        "debt_to_equity": "D/E", "free_cash_flow_cr": "FCF ₹ Cr",
        "revenue_cagr_5yr": "Revenue CAGR 5yr %", "pat_cagr_5yr": "PAT CAGR 5yr %",
        "operating_profit_margin_pct": "OPM %", "pe_ratio": "P/E", "pb_ratio": "P/B",
        "dividend_yield_pct": "Dividend Yield %", "interest_coverage": "ICR",
    }
)

if visible.empty:
    st.warning("No companies match the selected filters.")
else:
    st.dataframe(visible, use_container_width=True, hide_index=True, height=520)

csv_data = visible.to_csv(index=False).encode("utf-8-sig")
st.download_button(
    "Download visible results as CSV",
    data=csv_data,
    file_name="nifty100_screener_results.csv",
    mime="text/csv",
    disabled=visible.empty,
)
'''

PEERS_PY = PAGE_COMMON_HEADER + r'''
import pandas as pd
import plotly.graph_objects as go
import streamlit as st

from src.dashboard.utils.db import (
    get_peer_group_data,
    get_peer_group_names,
    get_peer_years,
)

st.title("Peer Comparison")
st.caption("Eight-metric radar chart and side-by-side KPI comparison for 11 peer groups")

groups = get_peer_group_names()
if not groups:
    st.warning("No peer groups are available.")
    st.stop()

group = st.selectbox("Peer group", groups, key="peer_group")
years = get_peer_years(group)
year_options = ["Latest available"] + years
year_choice = st.selectbox("Financial year", year_options, key="peer_year")
year = None if year_choice == "Latest available" else int(year_choice)
data = get_peer_group_data(group, year)

if data.empty:
    st.warning("No peer data are available for the selected group and year.")
    st.stop()

data = data.drop_duplicates(subset=["company_id"], keep="last").copy()
labels = {row.company_id: f"{row.company_name} — {row.company_id}" for row in data.itertuples()}
selected = st.selectbox(
    "Benchmark/selected company",
    data["company_id"].tolist(),
    format_func=lambda ticker: labels.get(ticker, ticker),
    key="peer_company",
)

metrics = {
    "ROE": "return_on_equity_pct",
    "ROCE": "return_on_capital_employed_pct",
    "Net Profit Margin": "net_profit_margin_pct",
    "D/E Score": "debt_to_equity",
    "FCF Score": "free_cash_flow_cr",
    "PAT CAGR 5yr": "pat_cagr_5yr",
    "Revenue CAGR 5yr": "revenue_cagr_5yr",
    "Composite Score": "composite_quality_score",
}

scores = pd.DataFrame(index=data.index)
for label, column in metrics.items():
    values = pd.to_numeric(data[column], errors="coerce")
    if values.notna().sum() <= 1:
        scores[label] = 50.0
    elif label == "D/E Score":
        scores[label] = values.rank(pct=True, ascending=False) * 100
    else:
        scores[label] = values.rank(pct=True, ascending=True) * 100
scores["company_id"] = data["company_id"].values

selected_row = scores[scores["company_id"] == selected]
if selected_row.empty:
    st.warning("Selected company has no peer metrics for this year.")
    st.stop()

theta = list(metrics.keys())
company_values = selected_row[theta].iloc[0].fillna(50).tolist()
peer_average = scores[theta].mean(skipna=True).fillna(50).tolist()

fig = go.Figure()
fig.add_trace(
    go.Scatterpolar(
        r=company_values + [company_values[0]],
        theta=theta + [theta[0]],
        fill="toself",
        name=selected,
    )
)
fig.add_trace(
    go.Scatterpolar(
        r=peer_average + [peer_average[0]],
        theta=theta + [theta[0]],
        mode="lines",
        line=dict(dash="dash"),
        name="Peer Group Average",
    )
)
fig.update_layout(
    title=f"{selected} vs {group} Average",
    polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
    showlegend=True,
    margin=dict(l=45, r=45, t=70, b=45),
)
st.plotly_chart(fig, use_container_width=True)

st.subheader("Peer KPI Table")
columns = [
    "company_id", "company_name", "is_benchmark", "return_on_equity_pct",
    "return_on_capital_employed_pct", "net_profit_margin_pct", "debt_to_equity",
    "free_cash_flow_cr", "pat_cagr_5yr", "revenue_cagr_5yr", "eps_cagr_5yr",
    "interest_coverage", "asset_turnover", "composite_quality_score",
]
columns = [column for column in columns if column in data.columns]
table = data[columns].rename(
    columns={
        "company_id": "Ticker", "company_name": "Company", "is_benchmark": "Benchmark",
        "return_on_equity_pct": "ROE %", "return_on_capital_employed_pct": "ROCE %",
        "net_profit_margin_pct": "NPM %", "debt_to_equity": "D/E",
        "free_cash_flow_cr": "FCF ₹ Cr", "pat_cagr_5yr": "PAT CAGR 5yr %",
        "revenue_cagr_5yr": "Revenue CAGR 5yr %", "eps_cagr_5yr": "EPS CAGR 5yr %",
        "interest_coverage": "ICR", "asset_turnover": "Asset Turnover",
        "composite_quality_score": "Composite Score",
    }
)

def highlight(row):
    return ["background-color: #ffe69c" if bool(row.get("Benchmark", False)) else "" for _ in row]

st.dataframe(table.style.apply(highlight, axis=1), use_container_width=True, hide_index=True)
'''

TRENDS_PY = PAGE_COMMON_HEADER + r'''
from functools import reduce

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st

from src.dashboard.utils.db import get_bs, get_cf, get_companies, get_pl, get_ratios
from src.dashboard.utils.formatting import clean_year

st.title("Trend Analysis")
st.caption("Overlay up to three metrics with ten-year YoY annotations")

companies = get_companies()
labels = {row.company_id: f"{row.company_name} — {row.company_id}" for row in companies.itertuples()}
ticker = st.selectbox(
    "Company",
    companies["company_id"].tolist(),
    format_func=lambda value: labels.get(value, value),
    key="trend_company",
)

frames = []
for frame in [clean_year(get_pl(ticker)), clean_year(get_bs(ticker)), clean_year(get_cf(ticker)), clean_year(get_ratios(ticker))]:
    if not frame.empty:
        frame = frame.sort_values("year").drop_duplicates("year", keep="last")
        frames.append(frame)
if not frames:
    st.warning("No trend data are available for this company.")
    st.stop()

prepared_frames: list[pd.DataFrame] = []

for frame in frames:
    if frame is None or frame.empty:
        continue

    current = frame.copy()

    if "year" not in current.columns:
        continue

    current["year"] = pd.to_numeric(
        current["year"],
        errors="coerce",
    )

    current = current.dropna(subset=["year"])
    current["year"] = current["year"].astype(int)

    current = current.drop(
        columns=["company_id", "id"],
        errors="ignore",
    )

    current = current.sort_values("year")
    current = current.drop_duplicates(
        subset=["year"],
        keep="last",
    )

    current = current.loc[
        :,
        ~current.columns.duplicated(),
    ]

    prepared_frames.append(
        current.set_index("year")
    )

if not prepared_frames:
    st.warning(
        "No financial history is available "
        "for the selected company."
    )
    st.stop()

data = pd.concat(
    prepared_frames,
    axis=1,
)

data = data.loc[
    :,
    ~data.columns.duplicated(keep="first"),
]

data = (
    data
    .reset_index()
    .sort_values("year")
    .reset_index(drop=True)
)
data = data.loc[:, ~data.columns.str.endswith("_dup")].sort_values("year").tail(10)

metric_map = {
    "Revenue": "sales",
    "Net Profit": "net_profit",
    "Operating Profit": "operating_profit",
    "ROE": "return_on_equity_pct",
    "ROCE": "return_on_capital_employed_pct",
    "Free Cash Flow": "free_cash_flow_cr",
    "Debt-to-Equity": "debt_to_equity",
    "Operating Profit Margin": "operating_profit_margin_pct",
    "Total Assets": "total_assets",
    "Borrowings": "borrowings",
}
available = [label for label, column in metric_map.items() if column in data.columns and pd.to_numeric(data[column], errors="coerce").notna().any()]
default = available[:2]
selected = st.multiselect("Metrics — select up to 3", available, default=default, key="trend_metrics")
if len(selected) > 3:
    st.warning("Select a maximum of three metrics.")
    st.stop()
if not selected:
    st.info("Select at least one metric.")
    st.stop()

normalise = st.checkbox("Normalise each metric to an index of 100", value=False)
fig = go.Figure()
for label in selected:
    column = metric_map[label]
    values = pd.to_numeric(data[column], errors="coerce")
    yoy = values.pct_change(fill_method=None) * 100
    plotted = values.copy()
    if normalise:
        valid = values.dropna()
        base = valid.iloc[0] if not valid.empty else np.nan
        plotted = values / base * 100 if pd.notna(base) and base != 0 else values
    text = ["" if pd.isna(value) else f"{value:+.1f}%" for value in yoy]
    fig.add_trace(
        go.Scatter(
            x=data["year"],
            y=plotted,
            mode="lines+markers+text",
            text=text,
            textposition="top center",
            name=label,
            connectgaps=False,
        )
    )
fig.update_layout(
    title=f"{ticker} — Latest {data['year'].nunique()} Available Years",
    xaxis_title="Financial Year",
    yaxis_title="Index (100 = first year)" if normalise else "Metric Value",
    hovermode="x unified",
    margin=dict(l=25, r=25, t=65, b=25),
)
st.plotly_chart(fig, use_container_width=True)
if data["year"].nunique() < 10:
    st.info(f"Only {data['year'].nunique()} years of combined data are available for this company.")
'''

SECTORS_PY = PAGE_COMMON_HEADER + r'''
import pandas as pd
import plotly.express as px
import streamlit as st

from src.dashboard.utils.db import get_sector_data, get_sector_names

st.title("Sector Analysis")
st.caption("Revenue vs ROE bubble map and sector median KPI profile")

sectors = get_sector_names()
sector = st.selectbox("Sector", sectors, key="sector_name")
year = st.sidebar.selectbox("Financial year", list(range(2024, 2018, -1)), key="sector_year")
data = get_sector_data(sector, year).copy()

if data.empty:
    st.warning("No sector data are available for this selection.")
    st.stop()

for column in ["sales", "return_on_equity_pct", "market_cap_crore", "free_cash_flow_cr", "pe_ratio"]:
    data[column] = pd.to_numeric(data[column], errors="coerce")

bubble = data.dropna(subset=["sales", "return_on_equity_pct", "market_cap_crore"])
bubble = bubble[bubble["market_cap_crore"] > 0]
if bubble.empty:
    st.info("The selected sector does not have enough data for the bubble chart.")
else:
    fig = px.scatter(
        bubble,
        x="sales",
        y="return_on_equity_pct",
        size="market_cap_crore",
        color="sub_sector",
        hover_name="company_name",
        hover_data={"company_id": True, "market_cap_crore": ":,.0f", "sales": ":,.0f"},
        size_max=65,
        title=f"{sector}: Revenue vs ROE ({year})",
        labels={"sales": "Revenue (₹ Cr)", "return_on_equity_pct": "ROE (%)"},
    )
    st.plotly_chart(fig, use_container_width=True)

if (data["market_cap_crore"] > 0).any():
    data["fcf_yield_pct"] = data["free_cash_flow_cr"] / data["market_cap_crore"] * 100
else:
    data["fcf_yield_pct"] = pd.NA

metrics = {
    "ROE %": "return_on_equity_pct",
    "ROCE %": "return_on_capital_employed_pct",
    "Net Profit Margin %": "net_profit_margin_pct",
    "Revenue CAGR 5yr %": "revenue_cagr_5yr",
    "PAT CAGR 5yr %": "pat_cagr_5yr",
    "Debt-to-Equity": "debt_to_equity",
    "P/E": "pe_ratio",
    "FCF Yield %": "fcf_yield_pct",
}
medians = []
for label, column in metrics.items():
    value = pd.to_numeric(data[column], errors="coerce").median()
    medians.append({"Metric": label, "Sector Median": value})
median_df = pd.DataFrame(medians).dropna(subset=["Sector Median"])
if median_df.empty:
    st.info("Sector median KPIs are unavailable.")
else:
    fig = px.bar(
        median_df,
        x="Metric",
        y="Sector Median",
        text_auto=".2f",
        title=f"{sector} Median KPIs — {year}",
    )
    fig.update_layout(xaxis_tickangle=-25)
    st.plotly_chart(fig, use_container_width=True)
'''

CAPITAL_PY = PAGE_COMMON_HEADER + r'''
import pandas as pd
import plotly.express as px
import streamlit as st

from src.dashboard.utils.db import get_capital_allocation_data

st.title("Capital Allocation Map")
st.caption("Latest capital allocation pattern for all 92 companies")

data = get_capital_allocation_data().copy()
if data.empty:
    st.warning("Capital allocation data are unavailable.")
    st.stop()

data["capital_allocation_pattern"] = data["capital_allocation_pattern"].fillna("Unclassified")
data["market_cap_crore"] = pd.to_numeric(data["market_cap_crore"], errors="coerce")
data["treemap_size"] = data["market_cap_crore"].where(data["market_cap_crore"] > 0, 1.0).fillna(1.0)

fig = px.treemap(
    data,
    path=["capital_allocation_pattern", "company_name"],
    values="treemap_size",
    hover_data=["company_id", "broad_sector", "free_cash_flow_cr", "cfo_quality_label"],
    title="Nifty 100 Capital Allocation Patterns",
)
fig.update_layout(margin=dict(l=5, r=5, t=55, b=5))
st.plotly_chart(fig, use_container_width=True)

patterns = sorted(data["capital_allocation_pattern"].dropna().unique().tolist())
selected_pattern = st.selectbox(
    "Select a pattern to view its companies",
    patterns,
    key="capital_pattern",
)
selected = data[data["capital_allocation_pattern"] == selected_pattern]
st.subheader(f"{selected_pattern}: {len(selected)} companies")
st.dataframe(
    selected[
        ["company_id", "company_name", "broad_sector", "free_cash_flow_cr", "cfo_quality_label", "market_cap_crore"]
    ].rename(
        columns={
            "company_id": "Ticker", "company_name": "Company", "broad_sector": "Sector",
            "free_cash_flow_cr": "FCF ₹ Cr", "cfo_quality_label": "CFO Quality",
            "market_cap_crore": "Market Cap ₹ Cr",
        }
    ),
    use_container_width=True,
    hide_index=True,
)
if "Unclassified" in patterns:
    count = int((data["capital_allocation_pattern"] == "Unclassified").sum())
    st.info(f"{count} companies do not have a latest capital-allocation classification.")
'''

REPORTS_PY = PAGE_COMMON_HEADER + r'''
from concurrent.futures import ThreadPoolExecutor

import pandas as pd
import requests
import streamlit as st

from src.dashboard.utils.db import get_companies, get_documents

st.title("Annual Reports")
st.caption("BSE annual-report PDF repository with cached availability checks")

companies = get_companies()
labels = {row.company_id: f"{row.company_name} — {row.company_id}" for row in companies.itertuples()}
ticker = st.selectbox(
    "Company",
    companies["company_id"].tolist(),
    format_func=lambda value: labels.get(value, value),
    key="report_company",
)
reports = get_documents(ticker)
if reports.empty:
    st.warning("No annual-report records are available for this company.")
    st.stop()

HEADERS = {"User-Agent": "Mozilla/5.0 N100-Analytics/1.0"}


def _check_one(url: str) -> tuple[str, int | None]:
    if not isinstance(url, str) or not url.strip():
        return "unavailable", None
    try:
        response = requests.head(url, allow_redirects=True, timeout=6, headers=HEADERS)
        status = response.status_code
        if status in (403, 405) or status >= 500:
            with requests.get(url, allow_redirects=True, timeout=7, headers=HEADERS, stream=True) as fallback:
                status = fallback.status_code
        if 200 <= status < 400:
            return "available", status
        if status == 404:
            return "unavailable", status
        return "unverified", status
    except requests.RequestException:
        return "unverified", None


@st.cache_data(ttl=3600, show_spinner=False)
def check_urls(urls: tuple[str, ...]) -> dict[str, tuple[str, int | None]]:
    with ThreadPoolExecutor(max_workers=min(8, max(1, len(urls)))) as executor:
        results = list(executor.map(_check_one, urls))
    return dict(zip(urls, results))

urls = tuple(reports["report_url"].fillna("").astype(str).tolist())
with st.spinner("Checking report links..."):
    availability = check_urls(urls)

for row in reports.itertuples():
    year = int(row.report_year) if pd.notna(row.report_year) else "Unknown year"
    url = str(row.report_url or "")
    state, status = availability.get(url, ("unverified", None))
    year_col, action_col = st.columns([1, 4])
    year_col.markdown(f"### {year}")
    if state == "available":
        action_col.link_button(f"Open {year} Annual Report", url, use_container_width=True)
    elif state == "unavailable":
        action_col.error(f"Report unavailable (HTTP {status or 404})")
    else:
        if url:
            action_col.markdown(f"[Open report without verification]({url})")
            action_col.warning(f"Availability could not be verified{f' (HTTP {status})' if status else ''}.")
        else:
            action_col.error("Report unavailable")
'''

VALUATION_PY = r'''"""Sprint 4 valuation engine: FCF yield and P/E valuation flags."""

from __future__ import annotations

from pathlib import Path
import sqlite3

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import get_settings

PROJECT_ROOT = Path(__file__).resolve().parents[2]
SETTINGS = get_settings()
DATABASE_PATH = Path(SETTINGS.database_path)
DEFAULT_OUTPUT_DIR = Path(SETTINGS.output_dir)


def _find_market_cap_excel() -> Path | None:
    candidates = [
        PROJECT_ROOT / "data" / "market_cap.xlsx",
        PROJECT_ROOT / "data" / "raw" / "market_cap.xlsx",
        PROJECT_ROOT / "data" / "raw" / "supporting datasets" / "market_cap.xlsx",
        PROJECT_ROOT / "supporting datasets" / "market_cap.xlsx",
        PROJECT_ROOT / "market_cap.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    for candidate in PROJECT_ROOT.rglob("market_cap.xlsx"):
        if ".venv" not in candidate.parts and ".git" not in candidate.parts:
            return candidate
    return None


def _load_market_cap(connection: sqlite3.Connection) -> pd.DataFrame:
    excel = _find_market_cap_excel()
    if excel is not None:
        data = pd.read_excel(excel)
        print(f"Market-cap source: {excel}")
        return data
    print(f"Market-cap source: SQLite table in {DATABASE_PATH}")
    return pd.read_sql_query("SELECT * FROM market_cap", connection)


def _style_excel(path: Path) -> None:
    workbook = load_workbook(path)
    sheet = workbook.active
    header_fill = PatternFill("solid", fgColor="1F4E78")
    caution_fill = PatternFill("solid", fgColor="F4CCCC")
    discount_fill = PatternFill("solid", fgColor="D9EAD3")
    fair_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    headers = {cell.value: cell.column for cell in sheet[1]}
    flag_col = headers.get("flag")
    if flag_col:
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=flag_col)
            if cell.value == "Caution":
                cell.fill = caution_fill
            elif cell.value == "Discount":
                cell.fill = discount_fill
            else:
                cell.fill = fair_fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in range(1, sheet.max_column + 1):
        max_length = max(len(str(sheet.cell(row=row, column=column).value or "")) for row in range(1, sheet.max_row + 1))
        sheet.column_dimensions[get_column_letter(column)].width = min(max(max_length + 2, 12), 34)
    workbook.save(path)


def build_valuation(output_dir: Path | str | None = None) -> tuple[pd.DataFrame, pd.DataFrame]:
    output_path = Path(output_dir) if output_dir is not None else DEFAULT_OUTPUT_DIR
    output_path.mkdir(parents=True, exist_ok=True)
    if not DATABASE_PATH.exists():
        raise FileNotFoundError(f"Database not found: {DATABASE_PATH}")

    with sqlite3.connect(DATABASE_PATH) as connection:
        companies = pd.read_sql_query(
            """
            SELECT c.company_id AS company_id, c.company_name, c.broad_sector AS sector
            FROM companies c
            ORDER BY c.company_id
            """,
            connection,
        )
        ratios = pd.read_sql_query(
            """
            SELECT company_id, CAST(year AS INTEGER) AS year, free_cash_flow_cr
            FROM financial_ratios
            """,
            connection,
        )
        market = _load_market_cap(connection)

    required_market_columns = {
        "company_id", "year", "market_cap_crore", "pe_ratio", "pb_ratio", "ev_ebitda"
    }
    missing = required_market_columns - set(market.columns)
    if missing:
        raise ValueError(f"market_cap data are missing columns: {sorted(missing)}")

    market = market.copy()
    market["company_id"] = market["company_id"].astype(str).str.strip().str.upper()
    market["year"] = pd.to_numeric(market["year"], errors="coerce")
    market = market.dropna(subset=["company_id", "year"])
    market["year"] = market["year"].astype(int)
    for column in ["market_cap_crore", "pe_ratio", "pb_ratio", "ev_ebitda"]:
        market[column] = pd.to_numeric(market[column], errors="coerce")

    ratios = ratios.copy()
    ratios["company_id"] = ratios["company_id"].astype(str).str.strip().str.upper()
    ratios["year"] = pd.to_numeric(ratios["year"], errors="coerce")
    ratios["free_cash_flow_cr"] = pd.to_numeric(ratios["free_cash_flow_cr"], errors="coerce")
    ratios = ratios.dropna(subset=["company_id", "year"])
    ratios["year"] = ratios["year"].astype(int)
    ratios = ratios.groupby(["company_id", "year"], as_index=False)["free_cash_flow_cr"].mean()

    latest_market = (
        market.sort_values(["company_id", "year"])
        .drop_duplicates("company_id", keep="last")
    )
    five_year_median = (
        market.sort_values(["company_id", "year"])
        .groupby("company_id", group_keys=False)
        .tail(5)
        .groupby("company_id")["pe_ratio"]
        .median()
        .rename("5yr_median_PE")
        .reset_index()
    )

    exact_fcf = latest_market[["company_id", "year"]].merge(
        ratios,
        on=["company_id", "year"],
        how="left",
    )
    latest_fcf = (
        ratios.sort_values(["company_id", "year"])
        .drop_duplicates("company_id", keep="last")
        [["company_id", "free_cash_flow_cr"]]
        .rename(columns={"free_cash_flow_cr": "fallback_fcf"})
    )
    exact_fcf = exact_fcf.merge(latest_fcf, on="company_id", how="left")
    exact_fcf["free_cash_flow_cr"] = exact_fcf["free_cash_flow_cr"].fillna(exact_fcf["fallback_fcf"])

    valuation = companies.merge(latest_market, on="company_id", how="left")
    valuation = valuation.merge(
        exact_fcf[["company_id", "free_cash_flow_cr"]],
        on="company_id",
        how="left",
    )
    valuation = valuation.merge(five_year_median, on="company_id", how="left")

    valuation["FCF_yield_pct"] = np.where(
        valuation["market_cap_crore"].gt(0),
        valuation["free_cash_flow_cr"] / valuation["market_cap_crore"] * 100,
        np.nan,
    )
    valuation["sector_median_pe"] = valuation.groupby("sector")["pe_ratio"].transform("median")
    valuation["PE_vs_sector_median_pct"] = np.where(
        valuation["sector_median_pe"].gt(0),
        (valuation["pe_ratio"] / valuation["sector_median_pe"] - 1) * 100,
        np.nan,
    )

    def assign_flag(row: pd.Series) -> str:
        pe = row["pe_ratio"]
        median = row["sector_median_pe"]
        if pd.isna(pe) or pd.isna(median) or median <= 0:
            return "Fair"
        if pe > median * 1.5:
            return "Caution"
        if pe < median * 0.7:
            return "Discount"
        return "Fair"

    valuation["flag"] = valuation.apply(assign_flag, axis=1)
    summary = valuation[
        [
            "company_id", "company_name", "sector", "pe_ratio", "pb_ratio", "ev_ebitda",
            "FCF_yield_pct", "5yr_median_PE", "PE_vs_sector_median_pct", "flag",
        ]
    ].rename(
        columns={
            "pe_ratio": "P/E",
            "pb_ratio": "P/B",
            "ev_ebitda": "EV/EBITDA",
        }
    )
    numeric = ["P/E", "P/B", "EV/EBITDA", "FCF_yield_pct", "5yr_median_PE", "PE_vs_sector_median_pct"]
    summary[numeric] = summary[numeric].round(2)

    if len(summary) != 92 or summary["company_id"].nunique() != 92:
        raise AssertionError(
            f"Expected 92 unique companies, found rows={len(summary)}, unique={summary['company_id'].nunique()}"
        )
    if not set(summary["flag"].dropna()).issubset({"Caution", "Discount", "Fair"}):
        raise AssertionError("Unexpected valuation flag value")

    flags = summary[summary["flag"].isin(["Caution", "Discount"])].copy()
    summary_file = output_path / "valuation_summary.xlsx"
    flags_file = output_path / "valuation_flags.csv"
    summary.to_excel(summary_file, index=False)
    _style_excel(summary_file)
    flags.to_csv(flags_file, index=False, encoding="utf-8-sig")

    print("Valuation module completed")
    print(f"Summary: {summary_file} ({len(summary)} rows)")
    print(f"Flags:   {flags_file} ({len(flags)} rows)")
    print(summary["flag"].value_counts().to_string())
    return summary, flags


if __name__ == "__main__":
    build_valuation()
'''

RUN_BAT = r'''@echo off
cd /d "%~dp0"
call .venv\Scripts\activate
set PYTHONPATH=%CD%
python -m src.analytics.valuation
if errorlevel 1 pause & exit /b 1
python -m streamlit run src\dashboard\app.py
'''

RETRO = r'''# Sprint 4 Retrospective — Dashboard & Valuation

## Scope Completed

- Eight-screen Streamlit dashboard
- Cached SQLite data-access layer with 600-second TTL
- Company profile, screener, peer, trend, sector, capital allocation and report screens
- Valuation summary and flagged-company CSV generation

## UX Decisions

- Missing numerical values display as `N/A`.
- Company selectors display both company name and NSE ticker.
- Partial histories display an information note instead of failing.
- The capital allocation treemap uses a pattern selector below the chart for reliable drill-down.
- Financial-sector companies are exempt from the normal D/E maximum filter, except in the Debt-Free preset.

## Data Edge Cases

- The database contains 1,073 financial-ratio records because source-year coverage is incomplete for some companies.
- Companies without stored pros/cons receive rule-based fallback observations.
- Annual-report links can be unavailable or blocked from automated verification; unverified links remain accessible.
- Companies without a latest capital-allocation label are shown as Unclassified.

## Performance Findings

Record warm-cache Company Profile load times here:

| Ticker | Load time | Pass under 3 sec |
|---|---:|---|
| TCS |  |  |
| HDFCBANK |  |  |
| RELIANCE |  |  |
| ITC |  |  |
| SUNPHARMA |  |  |

## QA Results

- 10 representative tickers tested: pending
- Extreme screener filters tested: pending
- CSV header validation: pending
- `valuation_summary.xlsx` row count: 92
- Team-lead demo/sign-off: pending
'''

README_SECTION = r'''

<!-- SPRINT4_DASHBOARD_START -->
## Sprint 4 — Streamlit Dashboard and Valuation

### Run the dashboard

```bash
set PYTHONPATH=%CD%
python -m streamlit run src/dashboard/app.py
```

The application opens at `http://localhost:8501`.

### Dashboard screens

1. **Home / Overview** — six market KPIs, sector donut chart and top-five quality companies.
2. **Company Profile** — company card, six KPIs, ten-year charts, pros and cons.
3. **Financial Screener** — ten sliders, six presets, live results and CSV export.
4. **Peer Comparison** — selected company versus peer-average radar and KPI table.
5. **Trend Analysis** — up to three overlaid metrics with YoY annotations.
6. **Sector Analysis** — revenue/ROE bubble chart and sector median KPIs.
7. **Capital Allocation Map** — treemap grouped by capital-allocation pattern.
8. **Annual Reports** — BSE PDF links with cached availability checks.

### Generate valuation outputs

```bash
set PYTHONPATH=%CD%
python -m src.analytics.valuation
```

Outputs:

- `output/valuation_summary.xlsx` — 92 companies with valuation multiples, FCF yield and flags.
- `output/valuation_flags.csv` — only Caution and Discount companies.

### Data coverage note

The validated database contains 92 companies and 1,073 financial-ratio records. The record count is below the original 1,100 target because some source company-year combinations are unavailable; no artificial rows are created.
<!-- SPRINT4_DASHBOARD_END -->
'''

FILES = {
    "src/dashboard/app.py": APP_PY,
    "src/dashboard/utils/db.py": DB_PY,
    "src/dashboard/utils/formatting.py": FORMATTING_PY,
    "src/dashboard/pages/01_home.py": HOME_PY,
    "src/dashboard/pages/02_profile.py": PROFILE_PY,
    "src/dashboard/pages/03_screener.py": SCREENER_PY,
    "src/dashboard/pages/04_peers.py": PEERS_PY,
    "src/dashboard/pages/05_trends.py": TRENDS_PY,
    "src/dashboard/pages/06_sectors.py": SECTORS_PY,
    "src/dashboard/pages/07_capital.py": CAPITAL_PY,
    "src/dashboard/pages/08_reports.py": REPORTS_PY,
    "src/analytics/valuation.py": VALUATION_PY,
    "run_sprint4.bat": RUN_BAT,
    "output/sprint4_retrospective.md": RETRO,
}


def update_requirements() -> None:
    path = ROOT / "requirements.txt"
    existing = path.read_text(encoding="utf-8") if path.exists() else ""
    required = [
        "streamlit>=1.30",
        "plotly>=5.18",
        "openpyxl>=3.1",
        "requests>=2.31",
    ]
    lines = existing.splitlines()
    lower = "\n".join(lines).lower()
    additions = [item for item in required if item.split(">=")[0].lower() not in lower]
    if additions:
        content = existing.rstrip() + "\n" + "\n".join(additions) + "\n"
        path.write_text(content.lstrip("\n"), encoding="utf-8")
        print("Updated: requirements.txt")


def update_readme() -> None:
    path = ROOT / "README.md"
    existing = path.read_text(encoding="utf-8") if path.exists() else "# N100 Financial Intelligence Platform\n"
    start = "<!-- SPRINT4_DASHBOARD_START -->"
    end = "<!-- SPRINT4_DASHBOARD_END -->"
    if start in existing and end in existing:
        prefix = existing.split(start, 1)[0].rstrip()
        suffix = existing.split(end, 1)[1].lstrip()
        updated = prefix + "\n" + README_SECTION.strip() + "\n"
        if suffix:
            updated += "\n" + suffix
    else:
        updated = existing.rstrip() + README_SECTION + "\n"
    path.write_text(updated, encoding="utf-8")
    print("Updated: README.md")


def compile_generated_files() -> None:
    failures = []
    for relative_path in FILES:
        if relative_path.endswith(".py"):
            try:
                py_compile.compile(str(ROOT / relative_path), doraise=True)
            except Exception as exc:  # noqa: BLE001
                failures.append(f"{relative_path}: {exc}")
    if failures:
        raise SystemExit("Generated code compilation failed:\n" + "\n".join(failures))
    print("Compilation check: PASS")


def main() -> None:
    require_project_root()
    for package_file in [
        "src/dashboard/__init__.py",
        "src/dashboard/pages/__init__.py",
        "src/dashboard/utils/__init__.py",
    ]:
        touch(package_file)
    for relative_path, content in FILES.items():
        backup_and_write(relative_path, content)
    update_requirements()
    update_readme()
    compile_generated_files()
    print("\nSprint 4 code generation completed successfully.")
    print("Next commands:")
    print("  python -m pip install -r requirements.txt")
    print("  set PYTHONPATH=%CD%")
    print("  python -m src.analytics.valuation")
    print("  python -m streamlit run src\\dashboard\\app.py")


if __name__ == "__main__":
    main()
