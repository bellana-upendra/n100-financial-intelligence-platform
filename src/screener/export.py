from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.screener.engine import (
    DEFAULT_CONFIG,
    apply_filter,
    load_config,
    run_preset,
)


PROJECT_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_DIR = PROJECT_ROOT / "output"
OUTPUT_FILE = OUTPUT_DIR / "screener_output.xlsx"


SHEET_ORDER = [
    "quality_compounder",
    "value_pick",
    "growth_accelerator",
    "dividend_champion",
    "debt_free_blue_chip",
    "turnaround_watch",
]


EXPORT_COLUMNS = [
    "company_id",
    "company_name",
    "broad_sector",
    "sub_sector",
    "year",
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "free_cash_flow_cr",
    "fcf_cagr_5yr_pct",
    "cfo_pat_ratio",
    "revenue_cagr_3yr",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct",
    "dividend_payout_pct",
    "market_cap_crore",
    "sales",
    "net_profit",
    "asset_turnover",
    "composite_quality_score",
    "sector_relative_score",
]


DISPLAY_NAMES = {
    "company_id": "Company ID",
    "company_name": "Company Name",
    "broad_sector": "Broad Sector",
    "sub_sector": "Sub Sector",
    "year": "Year",
    "return_on_equity_pct": "ROE (%)",
    "return_on_capital_employed_pct": "ROCE (%)",
    "net_profit_margin_pct": "Net Profit Margin (%)",
    "operating_profit_margin_pct": "Operating Profit Margin (%)",
    "debt_to_equity": "Debt / Equity",
    "interest_coverage": "Interest Coverage",
    "free_cash_flow_cr": "Free Cash Flow (Cr)",
    "fcf_cagr_5yr_pct": "FCF CAGR 5Y (%)",
    "cfo_pat_ratio": "CFO / PAT",
    "revenue_cagr_3yr": "Revenue CAGR 3Y (%)",
    "revenue_cagr_5yr": "Revenue CAGR 5Y (%)",
    "pat_cagr_5yr": "PAT CAGR 5Y (%)",
    "eps_cagr_5yr": "EPS CAGR 5Y (%)",
    "pe_ratio": "P/E",
    "pb_ratio": "P/B",
    "dividend_yield_pct": "Dividend Yield (%)",
    "dividend_payout_pct": "Dividend Payout (%)",
    "market_cap_crore": "Market Cap (Cr)",
    "sales": "Sales (Cr)",
    "net_profit": "Net Profit (Cr)",
    "asset_turnover": "Asset Turnover",
    "composite_quality_score": "Composite Score",
    "sector_relative_score": "Sector Relative Score",
}


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
SCORE_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN_GREY = Side(style="thin", color="D9E1F2")
CELL_BORDER = Border(
    left=THIN_GREY,
    right=THIN_GREY,
    top=THIN_GREY,
    bottom=THIN_GREY,
)


PERCENT_COLUMNS = {
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "fcf_cagr_5yr_pct",
    "revenue_cagr_3yr",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "dividend_yield_pct",
    "dividend_payout_pct",
}

RATIO_COLUMNS = {
    "debt_to_equity",
    "interest_coverage",
    "cfo_pat_ratio",
    "pe_ratio",
    "pb_ratio",
    "asset_turnover",
}

CRORE_COLUMNS = {
    "free_cash_flow_cr",
    "market_cap_crore",
    "sales",
    "net_profit",
}

SCORE_COLUMNS = {
    "composite_quality_score",
    "sector_relative_score",
}


def clean_text(value: Any) -> Any:
    """Remove embedded line breaks from exported text."""

    if isinstance(value, str):
        return " ".join(value.split())

    return value


def excel_safe_sheet_name(name: str) -> str:
    """Return an Excel-compatible sheet name."""

    invalid = set("[]:*?/\\")
    cleaned = "".join("_" if character in invalid else character for character in name)
    return cleaned[:31]


def prepare_export_data(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Select and clean the available export columns."""

    available_columns = [
        column for column in EXPORT_COLUMNS if column in dataframe.columns
    ]

    output = dataframe[available_columns].copy()

    for column in output.select_dtypes(include="object").columns:
        output[column] = output[column].map(clean_text)

    output = output.sort_values(
        ["composite_quality_score", "company_id"],
        ascending=[False, True],
        na_position="last",
    ).reset_index(drop=True)

    return output


def write_dataframe(
    worksheet,
    dataframe: pd.DataFrame,
) -> dict[str, int]:
    """Write one DataFrame and return source-column positions."""

    source_column_positions: dict[str, int] = {}

    for column_index, source_column in enumerate(dataframe.columns, start=1):
        source_column_positions[source_column] = column_index
        cell = worksheet.cell(
            row=1,
            column=column_index,
            value=DISPLAY_NAMES.get(source_column, source_column),
        )
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = CELL_BORDER

    for row_index, row_values in enumerate(
        dataframe.itertuples(index=False, name=None),
        start=2,
    ):
        for column_index, value in enumerate(row_values, start=1):
            if pd.isna(value):
                value = None

            cell = worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )
            cell.border = CELL_BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=False,
            )

    return source_column_positions


def apply_number_formats(
    worksheet,
    dataframe: pd.DataFrame,
    source_column_positions: dict[str, int],
) -> None:
    """Apply readable financial number formats."""

    last_row = len(dataframe) + 1

    for source_column, column_index in source_column_positions.items():
        number_format = None

        if source_column in PERCENT_COLUMNS:
            number_format = "0.00"
        elif source_column in RATIO_COLUMNS:
            number_format = "0.00"
        elif source_column in CRORE_COLUMNS:
            number_format = '#,##0.00'
        elif source_column in SCORE_COLUMNS:
            number_format = "0.00"
        elif source_column == "year":
            number_format = "0"

        if number_format is None:
            continue

        for row_index in range(2, last_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = number_format


def apply_threshold_colours(
    worksheet,
    result: pd.DataFrame,
    preset_filters: dict,
    config: dict,
    source_column_positions: dict[str, int],
) -> None:
    """Colour active preset metric cells green or red."""

    for metric_name, rule in preset_filters.items():
        metric_definition = config.get("metrics", {}).get(metric_name, {})
        source_column = metric_definition.get("column")

        if (
            source_column not in result.columns
            or source_column not in source_column_positions
        ):
            continue

        mask = apply_filter(
            result,
            metric_name,
            rule,
            config,
        )

        column_index = source_column_positions[source_column]

        for dataframe_index, passed in mask.items():
            excel_row = int(dataframe_index) + 2
            worksheet.cell(
                row=excel_row,
                column=column_index,
            ).fill = PASS_FILL if bool(passed) else FAIL_FILL


def apply_score_formatting(
    worksheet,
    dataframe: pd.DataFrame,
    source_column_positions: dict[str, int],
) -> None:
    """Highlight score columns and apply a 0–100 colour scale."""

    if dataframe.empty:
        return

    last_row = len(dataframe) + 1

    for source_column in SCORE_COLUMNS:
        column_index = source_column_positions.get(source_column)

        if column_index is None:
            continue

        for row_index in range(2, last_row + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).fill = SCORE_FILL

        column_letter = get_column_letter(column_index)
        worksheet.conditional_formatting.add(
            f"{column_letter}2:{column_letter}{last_row}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="F8696B",
                mid_type="num",
                mid_value=50,
                mid_color="FFEB84",
                end_type="num",
                end_value=100,
                end_color="63BE7B",
            ),
        )


def set_column_widths(
    worksheet,
    dataframe: pd.DataFrame,
) -> None:
    """Apply bounded column widths."""

    for column_index, source_column in enumerate(dataframe.columns, start=1):
        header = DISPLAY_NAMES.get(source_column, source_column)
        values = [
            "" if value is None else str(value)
            for value in dataframe[source_column].head(100)
        ]

        maximum_length = max(
            [len(str(header)), *[len(value) for value in values]],
            default=10,
        )

        if source_column == "company_name":
            width = min(max(maximum_length + 2, 24), 42)
        elif source_column in {"broad_sector", "sub_sector"}:
            width = min(max(maximum_length + 2, 16), 28)
        elif source_column == "company_id":
            width = min(max(maximum_length + 2, 12), 18)
        else:
            width = min(max(maximum_length + 2, 12), 22)

        worksheet.column_dimensions[
            get_column_letter(column_index)
        ].width = width


def format_worksheet(
    worksheet,
    dataframe: pd.DataFrame,
    preset_filters: dict,
    config: dict,
    source_column_positions: dict[str, int],
) -> None:
    """Apply all worksheet formatting."""

    last_row = len(dataframe) + 1
    last_column = len(dataframe.columns)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        f"A1:{get_column_letter(last_column)}{last_row}"
    )
    worksheet.row_dimensions[1].height = 36
    worksheet.sheet_view.showGridLines = False

    apply_number_formats(
        worksheet,
        dataframe,
        source_column_positions,
    )
    apply_threshold_colours(
        worksheet,
        dataframe,
        preset_filters,
        config,
        source_column_positions,
    )
    apply_score_formatting(
        worksheet,
        dataframe,
        source_column_positions,
    )
    set_column_widths(
        worksheet,
        dataframe,
    )


def create_screener_workbook(
    config_path: str | Path = DEFAULT_CONFIG,
    output_path: str | Path = OUTPUT_FILE,
) -> Path:
    """Create output/screener_output.xlsx with exactly six sheets."""

    config = load_config(config_path)
    output_path = Path(output_path)

    if not output_path.is_absolute():
        output_path = PROJECT_ROOT / output_path

    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    generated_counts: dict[str, int] = {}

    for preset_name in SHEET_ORDER:
        if preset_name not in config["presets"]:
            raise KeyError(
                f"Required preset missing from configuration: {preset_name}"
            )

        preset_definition = config["presets"][preset_name]
        result = run_preset(
            preset_name,
            config_path=config_path,
        )
        export_data = prepare_export_data(result)

        sheet_name = excel_safe_sheet_name(
            preset_definition["display_name"]
        )
        worksheet = workbook.create_sheet(sheet_name)

        positions = write_dataframe(
            worksheet,
            export_data,
        )

        format_worksheet(
            worksheet,
            export_data,
            preset_definition["filters"],
            config,
            positions,
        )

        generated_counts[sheet_name] = len(export_data)

    if len(workbook.sheetnames) != 6:
        raise ValueError(
            f"Expected 6 sheets, generated {len(workbook.sheetnames)}."
        )

    workbook.save(output_path)

    validation_workbook = load_workbook(
        output_path,
        read_only=True,
        data_only=False,
    )

    try:
        if len(validation_workbook.sheetnames) != 6:
            raise ValueError(
                "Saved workbook does not contain exactly 6 sheets."
            )
    finally:
        validation_workbook.close()

    print(f"Created: {output_path}")
    print(f"Sheet count: {len(generated_counts)}")

    for sheet_name, row_count in generated_counts.items():
        print(f"- {sheet_name}: {row_count} companies")

    return output_path


def main() -> None:
    create_screener_workbook()


if __name__ == "__main__":
    main()
