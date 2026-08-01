"""Sprint 5 Day 32: Capital allocation reporting.

Run from the project root:

    python -m src.reports.capital_allocation_report

Inputs:
    output/capital_allocation.csv
    output/cashflow_intelligence.xlsx
    configured SQLite database (for company names)

Outputs:
    output/capital_allocation_distribution.csv
    output/pattern_changes.csv
    output/cashflow_intelligence.xlsx  (updated in place)

Validation:
    - 92 unique companies
    - no duplicate company-year rows
    - no blank allocation patterns
    - exactly eight valid allocation patterns across the history
    - 'Insufficient Data' is treated as a status, not a ninth pattern
    - every cashflow company-year is present
    - additional allocation rows are reported but retained
    - one latest allocation label merged for every company
"""

from __future__ import annotations

import re
import sqlite3
from copy import copy
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook

from src.config import get_settings


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SETTINGS = get_settings()


def resolve_project_path(value: object) -> Path:
    """Resolve configured paths relative to the project root."""

    path = Path(str(value))
    return path if path.is_absolute() else PROJECT_ROOT / path


DATABASE_PATH = resolve_project_path(SETTINGS.database_path)
OUTPUT_DIR = resolve_project_path(SETTINGS.output_dir)

CAPITAL_ALLOCATION_PATH = OUTPUT_DIR / "capital_allocation.csv"
CASHFLOW_INTELLIGENCE_PATH = OUTPUT_DIR / "cashflow_intelligence.xlsx"
DISTRIBUTION_PATH = OUTPUT_DIR / "capital_allocation_distribution.csv"
PATTERN_CHANGES_PATH = OUTPUT_DIR / "pattern_changes.csv"

EXPECTED_COMPANY_COUNT = 92
EXPECTED_PATTERN_COUNT = 8

EXPECTED_VALID_PATTERNS = {
    "Cash Accumulator",
    "Distress Signal",
    "Growth Funded by Debt",
    "Liquidating Assets",
    "Mixed",
    "Pre-Revenue",
    "Reinvestor",
    "Shareholder Returns",
}

SPECIAL_STATUS_LABELS = {
    "Insufficient Data",
}

LABEL_COLUMN_CANDIDATES = (
    "capital_allocation_label",
    "capital_allocation_pattern",
    "pattern_label",
    "allocation_pattern",
    "pattern",
)

YEAR_COLUMN_CANDIDATES = (
    "financial_year",
    "year",
    "fy",
    "report_year",
)

COMPANY_COLUMN_CANDIDATES = (
    "company_id",
    "ticker",
    "symbol",
)


def normalise_column_name(value: object) -> str:
    """Convert a source column name to lowercase snake_case."""

    cleaned = re.sub(r"[^a-z0-9]+", "_", str(value).strip().lower())
    return cleaned.strip("_")


def normalise_company_id(value: object) -> str:
    """Strip whitespace and uppercase a company identifier."""

    if pd.isna(value):
        return ""
    return str(value).strip().upper()


def normalise_financial_year(value: object) -> int | None:
    """Convert values such as 2024, Mar-24, or 2024-03 into 2024."""

    if pd.isna(value):
        return None

    if isinstance(value, (int, float)):
        numeric = int(value)
        if 1900 <= numeric <= 2100:
            return numeric

    text = str(value).strip()

    four_digit = re.search(r"(?:19|20)\d{2}", text)
    if four_digit:
        return int(four_digit.group(0))

    two_digit = re.search(r"(?<!\d)(\d{2})(?!\d)", text)
    if two_digit:
        year = int(two_digit.group(1))
        return 2000 + year if year <= 79 else 1900 + year

    return None


def first_existing_column(
    columns: Iterable[str],
    candidates: Iterable[str],
) -> str | None:
    """Return the first candidate found in columns."""

    available = set(columns)

    for candidate in candidates:
        if candidate in available:
            return candidate

    return None


def clean_pattern_text(series: pd.Series) -> pd.Series:
    """Trim pattern text and convert common blank strings to missing."""

    cleaned = series.astype("string").str.strip()

    return cleaned.replace(
        {
            "": pd.NA,
            "nan": pd.NA,
            "NaN": pd.NA,
            "None": pd.NA,
            "<NA>": pd.NA,
        }
    )


def load_capital_allocation() -> tuple[pd.DataFrame, str]:
    """Load and standardise output/capital_allocation.csv."""

    if not CAPITAL_ALLOCATION_PATH.exists():
        raise FileNotFoundError(
            "Capital-allocation file was not found:\n"
            f"{CAPITAL_ALLOCATION_PATH}\n\n"
            "Run the existing capital-allocation generator first."
        )

    frame = pd.read_csv(CAPITAL_ALLOCATION_PATH)
    frame.columns = [
        normalise_column_name(column)
        for column in frame.columns
    ]

    company_column = first_existing_column(
        frame.columns,
        COMPANY_COLUMN_CANDIDATES,
    )
    year_column = first_existing_column(
        frame.columns,
        YEAR_COLUMN_CANDIDATES,
    )
    label_column = first_existing_column(
        frame.columns,
        LABEL_COLUMN_CANDIDATES,
    )

    missing_columns: list[str] = []

    if company_column is None:
        missing_columns.append("company_id/ticker")
    if year_column is None:
        missing_columns.append("financial_year")
    if label_column is None:
        missing_columns.append("capital-allocation label")

    if missing_columns:
        raise ValueError(
            "capital_allocation.csv is missing required fields: "
            + ", ".join(missing_columns)
            + f"\nColumns found: {frame.columns.tolist()}"
        )

    result = frame.copy()
    result["company_id"] = result[company_column].map(
        normalise_company_id
    )
    result["financial_year"] = result[year_column].map(
        normalise_financial_year
    )
    result["capital_allocation_label"] = clean_pattern_text(
        result[label_column]
    )

    result = result[
        result["company_id"] != ""
    ].copy()

    if result["financial_year"].isna().any():
        invalid = result.loc[
            result["financial_year"].isna(),
            [company_column, year_column],
        ].head(20)

        raise ValueError(
            "Some financial-year values could not be parsed:\n"
            + invalid.to_string(index=False)
        )

    result["financial_year"] = result["financial_year"].astype(int)

    result = result.sort_values(
        ["company_id", "financial_year"],
        kind="stable",
    ).reset_index(drop=True)

    return result, label_column


def find_internal_year_gaps(frame: pd.DataFrame) -> pd.DataFrame:
    """Find missing years between each company's first and latest year.

    Different companies may begin in different years. This check only flags
    missing years inside each company's own observed range.
    """

    gap_rows: list[dict[str, object]] = []

    for company_id, company_df in frame.groupby(
        "company_id",
        sort=True,
    ):
        observed = sorted(
            set(company_df["financial_year"].astype(int))
        )

        if len(observed) < 2:
            continue

        expected = set(
            range(
                observed[0],
                observed[-1] + 1,
            )
        )
        missing = sorted(expected - set(observed))

        for year in missing:
            gap_rows.append(
                {
                    "company_id": company_id,
                    "missing_financial_year": year,
                    "first_available_year": observed[0],
                    "latest_available_year": observed[-1],
                }
            )

    return pd.DataFrame(
        gap_rows,
        columns=[
            "company_id",
            "missing_financial_year",
            "first_available_year",
            "latest_available_year",
        ],
    )


def load_expected_company_years() -> pd.DataFrame:
    """Load expected company-year keys from the cashflow source table.

    Missing calendar years are not automatically errors because the original
    financial source can genuinely omit a year. Day 32 therefore validates
    capital_allocation.csv against the actual keys in the cashflow table.
    """

    if not DATABASE_PATH.exists():
        raise FileNotFoundError(
            f"Configured database was not found: {DATABASE_PATH}"
        )

    with sqlite3.connect(DATABASE_PATH) as connection:
        table_row = connection.execute(
            """
            SELECT 1
            FROM sqlite_master
            WHERE type = 'table'
              AND name = 'cashflow'
            LIMIT 1
            """
        ).fetchone()

        if table_row is None:
            raise RuntimeError(
                "The cashflow table was not found in the configured database."
            )

        source = pd.read_sql_query(
            "SELECT * FROM cashflow",
            connection,
        )

    source.columns = [
        normalise_column_name(column)
        for column in source.columns
    ]

    company_column = first_existing_column(
        source.columns,
        COMPANY_COLUMN_CANDIDATES,
    )
    year_column = first_existing_column(
        source.columns,
        YEAR_COLUMN_CANDIDATES,
    )

    if company_column is None or year_column is None:
        raise ValueError(
            "The cashflow table must contain company and year columns. "
            f"Columns found: {source.columns.tolist()}"
        )

    expected = pd.DataFrame(
        {
            "company_id": source[company_column].map(
                normalise_company_id
            ),
            "financial_year": source[year_column].map(
                normalise_financial_year
            ),
        }
    )

    expected = expected[
        (expected["company_id"] != "")
        & expected["financial_year"].notna()
    ].copy()

    expected["financial_year"] = expected["financial_year"].astype(int)

    return expected.drop_duplicates(
        ["company_id", "financial_year"]
    ).sort_values(
        ["company_id", "financial_year"],
        kind="stable",
    ).reset_index(drop=True)


def validate_capital_allocation(
    frame: pd.DataFrame,
    expected_company_years: pd.DataFrame,
) -> list[str]:
    """Validate company coverage, source years, duplicates, and patterns."""

    companies = int(frame["company_id"].nunique())
    rows = len(frame)
    duplicates = int(
        frame.duplicated(
            ["company_id", "financial_year"]
        ).sum()
    )
    blank_patterns = int(
        frame["capital_allocation_label"].isna().sum()
    )

    observed_labels = sorted(
        frame["capital_allocation_label"]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )

    valid_patterns = sorted(
        label
        for label in observed_labels
        if label not in SPECIAL_STATUS_LABELS
    )

    status_labels = sorted(
        label
        for label in observed_labels
        if label in SPECIAL_STATUS_LABELS
    )

    unknown_labels = sorted(
        set(observed_labels)
        - EXPECTED_VALID_PATTERNS
        - SPECIAL_STATUS_LABELS
    )

    actual_keys = frame[
        ["company_id", "financial_year"]
    ].drop_duplicates()

    expected_keys = expected_company_years[
        ["company_id", "financial_year"]
    ].drop_duplicates()

    coverage = expected_keys.merge(
        actual_keys,
        on=["company_id", "financial_year"],
        how="outer",
        indicator=True,
    )

    missing_expected_rows = coverage[
        coverage["_merge"] == "left_only"
    ][["company_id", "financial_year"]]

    unexpected_rows = coverage[
        coverage["_merge"] == "right_only"
    ][["company_id", "financial_year"]]

    calendar_gaps = find_internal_year_gaps(frame)

    print("Capital-allocation validation")
    print("=" * 60)
    print(f"Companies:               {companies}")
    print(f"Rows:                    {rows}")
    print(f"Expected source rows:    {len(expected_keys)}")
    print(f"Duplicates:              {duplicates}")
    print(f"Blank patterns:          {blank_patterns}")
    print(f"Valid patterns:          {len(valid_patterns)}")
    print(f"Status labels:           {len(status_labels)}")
    print(
        "Financial-year range:    "
        f"{frame['financial_year'].min()}–"
        f"{frame['financial_year'].max()}"
    )
    print(
        "Missing source rows:     "
        f"{len(missing_expected_rows)}"
    )
    print(
        "Unexpected source rows:  "
        f"{len(unexpected_rows)}"
    )
    print(
        "Calendar gaps in source: "
        f"{len(calendar_gaps)} (informational)"
    )

    print()
    print("Eight valid patterns found:")

    for pattern in valid_patterns:
        print(f"  - {pattern}")

    if status_labels:
        print()
        print("Special status labels:")

        for label in status_labels:
            count = int(
                (
                    frame["capital_allocation_label"]
                    == label
                ).sum()
            )
            print(f"  - {label}: {count} rows")

    if companies != EXPECTED_COMPANY_COUNT:
        raise RuntimeError(
            f"Expected {EXPECTED_COMPANY_COUNT} companies, "
            f"but found {companies}."
        )

    if duplicates:
        duplicate_sample = frame.loc[
            frame.duplicated(
                ["company_id", "financial_year"],
                keep=False,
            ),
            [
                "company_id",
                "financial_year",
                "capital_allocation_label",
            ],
        ].head(20)

        raise RuntimeError(
            "Duplicate company-year rows were found:\n"
            + duplicate_sample.to_string(index=False)
        )

    if blank_patterns:
        blank_sample = frame.loc[
            frame["capital_allocation_label"].isna(),
            ["company_id", "financial_year"],
        ].head(20)

        raise RuntimeError(
            "Rows with blank allocation patterns were found:\n"
            + blank_sample.to_string(index=False)
        )

    if unknown_labels:
        raise RuntimeError(
            "Unexpected allocation labels were found: "
            f"{unknown_labels}"
        )

    if set(valid_patterns) != EXPECTED_VALID_PATTERNS:
        missing_patterns = sorted(
            EXPECTED_VALID_PATTERNS - set(valid_patterns)
        )
        extra_patterns = sorted(
            set(valid_patterns) - EXPECTED_VALID_PATTERNS
        )

        raise RuntimeError(
            "The eight valid patterns do not match the expected set. "
            f"Missing={missing_patterns}; Extra={extra_patterns}"
        )

    if len(valid_patterns) != EXPECTED_PATTERN_COUNT:
        raise RuntimeError(
            f"Expected {EXPECTED_PATTERN_COUNT} valid patterns, "
            f"but found {len(valid_patterns)}."
        )

    if not missing_expected_rows.empty:
        raise RuntimeError(
            "capital_allocation.csv is missing company-year rows that "
            "exist in the cashflow source table:\n"
            + missing_expected_rows.head(30).to_string(index=False)
        )

    if not unexpected_rows.empty:
        print()
        print(
            "Additional allocation rows not present in the raw cashflow "
            "table were found. They are retained because the allocation "
            "file is the reporting source for Day 32:"
        )
        print(unexpected_rows.head(30).to_string(index=False))

    # Include special statuses in the distribution so all latest companies
    # remain counted, while validating only the eight true patterns.
    return observed_labels


def select_latest_rows(frame: pd.DataFrame) -> pd.DataFrame:
    """Select exactly one latest financial-year row per company."""

    latest = (
        frame.sort_values(
            ["company_id", "financial_year"],
            kind="stable",
        )
        .groupby(
            "company_id",
            as_index=False,
            sort=False,
        )
        .tail(1)
        .copy()
    )

    latest = latest[
        [
            "company_id",
            "financial_year",
            "capital_allocation_label",
        ]
    ].sort_values(
        "company_id",
        kind="stable",
    ).reset_index(drop=True)

    if len(latest) != EXPECTED_COMPANY_COUNT:
        raise RuntimeError(
            "Latest-year selection did not produce "
            f"{EXPECTED_COMPANY_COUNT} companies. Found {len(latest)}."
        )

    return latest


def create_distribution(
    latest: pd.DataFrame,
    all_patterns: list[str],
) -> pd.DataFrame:
    """Create the latest-year pattern distribution."""

    counts = (
        latest["capital_allocation_label"]
        .value_counts()
        .reindex(all_patterns, fill_value=0)
    )

    total = len(latest)

    distribution = pd.DataFrame(
        {
            "capital_allocation_label": counts.index,
            "company_count": counts.values,
            "percentage": (
                counts.values / total * 100.0
            ).round(2),
        }
    )

    distribution = distribution.sort_values(
        ["company_count", "capital_allocation_label"],
        ascending=[False, True],
        kind="stable",
    ).reset_index(drop=True)

    return distribution


def load_company_names() -> pd.DataFrame:
    """Load company names from the configured SQLite database."""

    if not DATABASE_PATH.exists():
        raise FileNotFoundError(
            f"Configured database was not found: {DATABASE_PATH}"
        )

    with sqlite3.connect(DATABASE_PATH) as connection:
        companies = pd.read_sql_query(
            """
            SELECT company_id, company_name
            FROM companies
            """,
            connection,
        )

    companies["company_id"] = companies["company_id"].map(
        normalise_company_id
    )
    companies["company_name"] = (
        companies["company_name"]
        .astype("string")
        .str.strip()
        .fillna(companies["company_id"])
    )

    return companies.drop_duplicates(
        "company_id",
        keep="last",
    )


def create_pattern_changes(
    frame: pd.DataFrame,
    companies: pd.DataFrame,
) -> pd.DataFrame:
    """Create one row whenever a company's pattern changes year over year."""

    ordered = frame[
        [
            "company_id",
            "financial_year",
            "capital_allocation_label",
        ]
    ].copy()

    ordered = ordered.sort_values(
        ["company_id", "financial_year"],
        kind="stable",
    )

    ordered["previous_year"] = ordered.groupby(
        "company_id",
        sort=False,
    )["financial_year"].shift(1)

    ordered["previous_pattern"] = ordered.groupby(
        "company_id",
        sort=False,
    )["capital_allocation_label"].shift(1)

    changed = ordered[
        ordered["previous_pattern"].notna()
        & (
            ordered["capital_allocation_label"]
            != ordered["previous_pattern"]
        )
    ].copy()

    changed = changed.rename(
        columns={
            "financial_year": "current_year",
            "capital_allocation_label": "current_pattern",
        }
    )

    changed["previous_year"] = changed["previous_year"].astype(int)
    changed["current_year"] = changed["current_year"].astype(int)

    changed = changed.merge(
        companies,
        on="company_id",
        how="left",
        validate="many_to_one",
    )

    changed["ticker"] = changed["company_id"]
    changed["company_name"] = changed["company_name"].fillna(
        changed["company_id"]
    )

    changed["change_description"] = (
        "Moved from "
        + changed["previous_pattern"].astype(str)
        + " to "
        + changed["current_pattern"].astype(str)
    )

    return changed[
        [
            "company_id",
            "ticker",
            "company_name",
            "previous_year",
            "current_year",
            "previous_pattern",
            "current_pattern",
            "change_description",
        ]
    ].sort_values(
        ["company_id", "current_year"],
        kind="stable",
    ).reset_index(drop=True)


def update_cashflow_workbook(
    latest: pd.DataFrame,
) -> None:
    """Merge latest allocation labels into the existing Excel workbook.

    openpyxl is used instead of rewriting through pandas so the Day 31 workbook
    formatting, widths, filters, and frozen panes are preserved.
    """

    if not CASHFLOW_INTELLIGENCE_PATH.exists():
        raise FileNotFoundError(
            "Cash Flow Intelligence workbook was not found:\n"
            f"{CASHFLOW_INTELLIGENCE_PATH}\n\n"
            "Complete Day 31 before running Day 32."
        )

    latest_map = dict(
        zip(
            latest["company_id"],
            latest["capital_allocation_label"],
        )
    )

    workbook = load_workbook(CASHFLOW_INTELLIGENCE_PATH)
    worksheet = workbook[
        "Cash Flow Intelligence"
    ] if "Cash Flow Intelligence" in workbook.sheetnames else workbook.active

    headers = {
        normalise_column_name(cell.value): cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }

    company_column = first_existing_column(
        headers.keys(),
        COMPANY_COLUMN_CANDIDATES,
    )

    if company_column is None:
        raise ValueError(
            "The Cash Flow Intelligence workbook has no company_id column."
        )

    company_column_index = headers[company_column]

    allocation_header = "capital_allocation_label"

    if allocation_header in headers:
        allocation_column_index = headers[allocation_header]
    else:
        allocation_column_index = worksheet.max_column + 1
        header_cell = worksheet.cell(
            row=1,
            column=allocation_column_index,
            value=allocation_header,
        )

        if worksheet.max_column > 1:
            source_cell = worksheet.cell(
                row=1,
                column=allocation_column_index - 1,
            )
            header_cell.font = copy(source_cell.font)
            header_cell.fill = copy(source_cell.fill)
            header_cell.border = copy(source_cell.border)
            header_cell.alignment = copy(source_cell.alignment)
            header_cell.number_format = source_cell.number_format

        worksheet.column_dimensions[
            header_cell.column_letter
        ].width = 31

    workbook_ids: set[str] = set()

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        company_id = normalise_company_id(
            worksheet.cell(
                row=row_number,
                column=company_column_index,
            ).value
        )

        if not company_id:
            continue

        workbook_ids.add(company_id)

        worksheet.cell(
            row=row_number,
            column=allocation_column_index,
            value=latest_map.get(company_id),
        )

    missing_in_workbook = sorted(
        set(latest_map) - workbook_ids
    )
    missing_labels = sorted(
        company_id
        for company_id in workbook_ids
        if company_id not in latest_map
    )

    if missing_in_workbook or missing_labels:
        raise RuntimeError(
            "Excel merge coverage failed. "
            f"Missing workbook companies={missing_in_workbook}; "
            f"Missing latest labels={missing_labels}"
        )

    worksheet.auto_filter.ref = worksheet.dimensions
    workbook.save(CASHFLOW_INTELLIGENCE_PATH)


def validate_updated_workbook(
    latest: pd.DataFrame,
) -> None:
    """Validate the updated Excel workbook after saving."""

    workbook_data = pd.read_excel(CASHFLOW_INTELLIGENCE_PATH)
    workbook_data.columns = [
        normalise_column_name(column)
        for column in workbook_data.columns
    ]

    required = {
        "company_id",
        "capital_allocation_label",
    }

    if not required.issubset(workbook_data.columns):
        raise RuntimeError(
            "Updated workbook is missing required columns. "
            f"Columns found: {workbook_data.columns.tolist()}"
        )

    workbook_data["company_id"] = workbook_data["company_id"].map(
        normalise_company_id
    )
    workbook_data["capital_allocation_label"] = clean_pattern_text(
        workbook_data["capital_allocation_label"]
    )

    if len(workbook_data) != EXPECTED_COMPANY_COUNT:
        raise RuntimeError(
            "Updated workbook should contain "
            f"{EXPECTED_COMPANY_COUNT} rows, found {len(workbook_data)}."
        )

    duplicates = int(
        workbook_data["company_id"].duplicated().sum()
    )

    if duplicates:
        raise RuntimeError(
            f"Updated workbook contains {duplicates} duplicate companies."
        )

    blanks = int(
        workbook_data["capital_allocation_label"].isna().sum()
    )

    if blanks:
        raise RuntimeError(
            f"Updated workbook contains {blanks} blank allocation labels."
        )

    expected = latest.set_index("company_id")[
        "capital_allocation_label"
    ]
    actual = workbook_data.set_index("company_id")[
        "capital_allocation_label"
    ]

    comparison = pd.DataFrame(
        {
            "expected": expected,
            "actual": actual,
        }
    )

    mismatches = comparison[
        comparison["expected"] != comparison["actual"]
    ]

    if not mismatches.empty:
        raise RuntimeError(
            "Latest labels were not merged correctly:\n"
            + mismatches.head(20).to_string()
        )


def run_report() -> tuple[pd.DataFrame, pd.DataFrame]:
    """Run the complete Day 32 reporting workflow."""

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    capital_allocation, source_label_column = load_capital_allocation()
    expected_company_years = load_expected_company_years()
    distribution_labels = validate_capital_allocation(
        capital_allocation,
        expected_company_years,
    )

    latest = select_latest_rows(capital_allocation)
    companies = load_company_names()

    distribution = create_distribution(
        latest,
        distribution_labels,
    )
    changes = create_pattern_changes(
        capital_allocation,
        companies,
    )

    distribution.to_csv(
        DISTRIBUTION_PATH,
        index=False,
    )
    changes.to_csv(
        PATTERN_CHANGES_PATH,
        index=False,
    )

    update_cashflow_workbook(latest)
    validate_updated_workbook(latest)

    print()
    print("Day 32 capital-allocation reporting completed")
    print("=" * 60)
    print(f"Source label column:      {source_label_column}")
    print(f"Latest company rows:      {len(latest)}")
    print(f"Distribution rows:        {len(distribution)}")
    print(f"Pattern-change rows:      {len(changes)}")
    print(f"Updated workbook:         {CASHFLOW_INTELLIGENCE_PATH}")
    print(f"Distribution output:      {DISTRIBUTION_PATH}")
    print(f"Pattern changes output:   {PATTERN_CHANGES_PATH}")
    print("Coverage check: PASS — all 92 companies were updated.")

    print()
    print("Latest-year distribution:")
    print(distribution.to_string(index=False))

    return distribution, changes


def main() -> None:
    """Command-line entry point."""

    run_report()


if __name__ == "__main__":
    main()
