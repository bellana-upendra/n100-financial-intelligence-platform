from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.analytics.peer import (
    build_peer_dataset,
    calculate_peer_percentiles,
)
from src.screener.engine import (
    DEFAULT_CONFIG,
    load_config,
    load_financial_data,
)


PROJECT_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_FILE = PROJECT_ROOT / "output" / "peer_comparison.xlsx"


RAW_METRICS = [
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "free_cash_flow_cr",
    "fcf_cagr_5yr_pct",
    "cfo_pat_ratio",
    "revenue_cagr_3yr",
    "revenue_cagr_5yr",
    "pat_cagr_3yr",
    "pat_cagr_5yr",
    "eps_cagr_3yr",
    "eps_cagr_5yr",
    "asset_turnover",
    "dividend_yield_pct",
    "pe_ratio",
    "pb_ratio",
    "composite_quality_score",
]


PERCENTILE_COLUMNS = [
    "roe_percentile",
    "roce_percentile",
    "net_profit_margin_percentile",
    "debt_to_equity_percentile",
    "free_cash_flow_percentile",
    "pat_cagr_5yr_percentile",
    "revenue_cagr_5yr_percentile",
    "eps_cagr_5yr_percentile",
    "interest_coverage_percentile",
    "asset_turnover_percentile",
]


ID_COLUMNS = [
    "company_id",
    "company_name",
    "broad_sector",
    "sub_sector",
    "year",
    "is_benchmark",
]


DISPLAY_NAMES = {
    "company_id": "Company ID",
    "company_name": "Company Name",
    "broad_sector": "Broad Sector",
    "sub_sector": "Sub Sector",
    "year": "Year",
    "is_benchmark": "Benchmark",
    "return_on_equity_pct": "ROE (%)",
    "return_on_capital_employed_pct": "ROCE (%)",
    "net_profit_margin_pct": "Net Profit Margin (%)",
    "operating_profit_margin_pct": "Operating Profit Margin (%)",
    "debt_to_equity": "Debt / Equity",
    "interest_coverage": "Interest Coverage",
    "free_cash_flow_cr": "Free Cash Flow (Cr)",
    "fcf_cagr_5yr_pct": "FCF CAGR 5Y (%)",
    "cfo_pat_ratio": "CFO / PAT",
    "revenue_cagr_3yr": "Revenue CAGR 3Y (%)",
    "revenue_cagr_5yr": "Revenue CAGR 5Y (%)",
    "pat_cagr_3yr": "PAT CAGR 3Y (%)",
    "pat_cagr_5yr": "PAT CAGR 5Y (%)",
    "eps_cagr_3yr": "EPS CAGR 3Y (%)",
    "eps_cagr_5yr": "EPS CAGR 5Y (%)",
    "asset_turnover": "Asset Turnover",
    "dividend_yield_pct": "Dividend Yield (%)",
    "pe_ratio": "P/E",
    "pb_ratio": "P/B",
    "composite_quality_score": "Composite Score",
    "roe_percentile": "ROE Percentile",
    "roce_percentile": "ROCE Percentile",
    "net_profit_margin_percentile": "NPM Percentile",
    "debt_to_equity_percentile": "D/E Percentile",
    "free_cash_flow_percentile": "FCF Percentile",
    "pat_cagr_5yr_percentile": "PAT CAGR 5Y Percentile",
    "revenue_cagr_5yr_percentile": "Revenue CAGR 5Y Percentile",
    "eps_cagr_5yr_percentile": "EPS CAGR 5Y Percentile",
    "interest_coverage_percentile": "Interest Coverage Percentile",
    "asset_turnover_percentile": "Asset Turnover Percentile",
}


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BENCHMARK_FILL = PatternFill("solid", fgColor="FFD966")
MEDIAN_FILL = PatternFill("solid", fgColor="D9EAD3")
HIGH_FILL = PatternFill("solid", fgColor="C6EFCE")
MID_FILL = PatternFill("solid", fgColor="FFEB9C")
LOW_FILL = PatternFill("solid", fgColor="FFC7CE")
THIN_SIDE = Side(style="thin", color="D9E1F2")
CELL_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)


PERCENT_STYLE_COLUMNS = {
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "fcf_cagr_5yr_pct",
    "revenue_cagr_3yr",
    "revenue_cagr_5yr",
    "pat_cagr_3yr",
    "pat_cagr_5yr",
    "eps_cagr_3yr",
    "eps_cagr_5yr",
    "dividend_yield_pct",
}


RATIO_STYLE_COLUMNS = {
    "debt_to_equity",
    "interest_coverage",
    "cfo_pat_ratio",
    "asset_turnover",
    "pe_ratio",
    "pb_ratio",
    "composite_quality_score",
}


CRORE_STYLE_COLUMNS = {
    "free_cash_flow_cr",
}


def clean_text(value: Any) -> Any:
    """Remove line breaks from exported text."""

    if isinstance(value, str):
        return " ".join(value.split())

    return value


def excel_safe_sheet_name(name: str) -> str:
    """Return a valid unique-ready Excel sheet name."""

    invalid = set("[]:*?/\\")
    cleaned = "".join(
        "_" if character in invalid else character
        for character in str(name)
    )
    return cleaned[:31]


def build_export_dataset(
    config_path: str | Path = DEFAULT_CONFIG,
) -> pd.DataFrame:
    """Build one wide dataset containing raw metrics and percentiles."""

    config = load_config(config_path)
    all_financial_data = load_financial_data(config)

    peer_data, _ = build_peer_dataset(config_path)
    wide_peer_data, _ = calculate_peer_percentiles(peer_data)

    required_financial_columns = [
        "company_id",
        *RAW_METRICS,
    ]

    missing_financial_columns = [
        column
        for column in required_financial_columns
        if column not in all_financial_data.columns
    ]

    if missing_financial_columns:
        raise KeyError(
            "Missing financial export columns: "
            + ", ".join(missing_financial_columns)
        )

    enrichment_columns = [
        column
        for column in RAW_METRICS
        if column not in wide_peer_data.columns
    ]

    if enrichment_columns:
        wide_peer_data = wide_peer_data.merge(
            all_financial_data[
                ["company_id", *enrichment_columns]
            ],
            on="company_id",
            how="left",
            validate="many_to_one",
        )

    missing_percentiles = [
        column
        for column in PERCENTILE_COLUMNS
        if column not in wide_peer_data.columns
    ]

    if missing_percentiles:
        raise KeyError(
            "Missing percentile columns: "
            + ", ".join(missing_percentiles)
        )

    export_columns = [
        "peer_group_name",
        *ID_COLUMNS,
        *RAW_METRICS,
        *PERCENTILE_COLUMNS,
    ]

    export_data = wide_peer_data[export_columns].copy()

    for column in export_data.columns:
        if (
            pd.api.types.is_object_dtype(export_data[column])
            or pd.api.types.is_string_dtype(export_data[column])
        ):
            export_data[column] = export_data[column].map(clean_text)

    export_data["is_benchmark"] = (
        export_data["is_benchmark"]
        .fillna(0)
        .astype(int)
    )

    return export_data


def add_median_row(
    group_data: pd.DataFrame,
) -> pd.DataFrame:
    """Append a peer-group median row."""

    median_row: dict[str, Any] = {
        column: None
        for column in group_data.columns
    }

    median_row["company_id"] = "Peer Group Median"
    median_row["company_name"] = ""
    median_row["broad_sector"] = ""
    median_row["sub_sector"] = ""
    median_row["year"] = int(group_data["year"].max())
    median_row["is_benchmark"] = None

    for column in [
        *RAW_METRICS,
        *PERCENTILE_COLUMNS,
    ]:
        median_row[column] = pd.to_numeric(
            group_data[column],
            errors="coerce",
        ).median()

    return pd.concat(
        [
            group_data,
            pd.DataFrame([median_row]),
        ],
        ignore_index=True,
    )


def write_sheet(
    worksheet,
    group_data: pd.DataFrame,
) -> None:
    """Write and format one peer-group worksheet."""

    display_columns = [
        *ID_COLUMNS,
        *RAW_METRICS,
        *PERCENTILE_COLUMNS,
    ]

    data_with_median = add_median_row(
        group_data[display_columns].copy()
    )

    column_positions: dict[str, int] = {}

    for column_index, source_column in enumerate(
        display_columns,
        start=1,
    ):
        column_positions[source_column] = column_index
        header_cell = worksheet.cell(
            row=1,
            column=column_index,
            value=DISPLAY_NAMES.get(
                source_column,
                source_column,
            ),
        )
        header_cell.fill = HEADER_FILL
        header_cell.font = HEADER_FONT
        header_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        header_cell.border = CELL_BORDER

    for row_index, row_values in enumerate(
        data_with_median.itertuples(
            index=False,
            name=None,
        ),
        start=2,
    ):
        for column_index, value in enumerate(
            row_values,
            start=1,
        ):
            if pd.isna(value):
                value = None

            cell = worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )
            cell.border = CELL_BORDER
            cell.alignment = Alignment(
                vertical="center",
            )

    data_row_count = len(group_data)
    first_data_row = 2
    last_data_row = data_row_count + 1
    median_row_number = data_row_count + 2
    last_column = len(display_columns)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        f"A1:{get_column_letter(last_column)}"
        f"{last_data_row}"
    )
    worksheet.row_dimensions[1].height = 36
    worksheet.sheet_view.showGridLines = False

    benchmark_column = column_positions["is_benchmark"]

    for row_number in range(
        first_data_row,
        last_data_row + 1,
    ):
        benchmark_value = worksheet.cell(
            row=row_number,
            column=benchmark_column,
        ).value

        if benchmark_value == 1:
            for column_index in range(
                1,
                last_column + 1,
            ):
                worksheet.cell(
                    row=row_number,
                    column=column_index,
                ).fill = BENCHMARK_FILL

    for percentile_column in PERCENTILE_COLUMNS:
        column_index = column_positions[
            percentile_column
        ]

        for row_number in range(
            first_data_row,
            last_data_row + 1,
        ):
            cell = worksheet.cell(
                row=row_number,
                column=column_index,
            )

            if cell.value is None:
                continue

            value = float(cell.value)

            if value >= 0.75:
                cell.fill = HIGH_FILL
            elif value <= 0.25:
                cell.fill = LOW_FILL
            else:
                cell.fill = MID_FILL

            cell.number_format = "0%"

    for column_index in range(
        1,
        last_column + 1,
    ):
        median_cell = worksheet.cell(
            row=median_row_number,
            column=column_index,
        )
        median_cell.fill = MEDIAN_FILL
        median_cell.font = Font(bold=True)

    for source_column, column_index in column_positions.items():
        if source_column in PERCENT_STYLE_COLUMNS:
            number_format = "0.00"
        elif source_column in RATIO_STYLE_COLUMNS:
            number_format = "0.00"
        elif source_column in CRORE_STYLE_COLUMNS:
            number_format = "#,##0.00"
        elif source_column in PERCENTILE_COLUMNS:
            number_format = "0%"
        elif source_column == "year":
            number_format = "0"
        else:
            number_format = None

        if number_format:
            for row_number in range(
                first_data_row,
                median_row_number + 1,
            ):
                worksheet.cell(
                    row=row_number,
                    column=column_index,
                ).number_format = number_format

        values = [
            DISPLAY_NAMES.get(
                source_column,
                source_column,
            )
        ]

        values.extend(
            ""
            if value is None
            else str(value)
            for value in data_with_median[
                source_column
            ].head(100)
        )

        maximum_length = max(
            len(value)
            for value in values
        )

        if source_column == "company_name":
            width = min(
                max(maximum_length + 2, 24),
                42,
            )
        elif source_column in {
            "broad_sector",
            "sub_sector",
        }:
            width = min(
                max(maximum_length + 2, 16),
                28,
            )
        elif source_column == "company_id":
            width = min(
                max(maximum_length + 2, 16),
                22,
            )
        else:
            width = min(
                max(maximum_length + 2, 12),
                24,
            )

        worksheet.column_dimensions[
            get_column_letter(column_index)
        ].width = width


def create_peer_comparison_workbook(
    config_path: str | Path = DEFAULT_CONFIG,
    output_path: str | Path = OUTPUT_FILE,
) -> Path:
    """Generate output/peer_comparison.xlsx with exactly 11 sheets."""

    output_path = Path(output_path)

    if not output_path.is_absolute():
        output_path = PROJECT_ROOT / output_path

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    export_data = build_export_dataset(
        config_path
    )

    peer_groups = sorted(
        export_data["peer_group_name"]
        .dropna()
        .unique()
        .tolist()
    )

    if len(peer_groups) != 11:
        raise AssertionError(
            f"Expected 11 peer groups, found {len(peer_groups)}."
        )

    workbook = Workbook()
    workbook.remove(workbook.active)

    used_sheet_names: set[str] = set()
    row_counts: dict[str, int] = {}

    for peer_group_name in peer_groups:
        group_data = export_data[
            export_data["peer_group_name"].eq(
                peer_group_name
            )
        ].copy()

        group_data = group_data.sort_values(
            [
                "composite_quality_score",
                "company_id",
            ],
            ascending=[False, True],
            na_position="last",
        ).reset_index(drop=True)

        sheet_name = excel_safe_sheet_name(
            peer_group_name
        )

        base_name = sheet_name
        suffix = 1

        while sheet_name in used_sheet_names:
            suffix_text = f"_{suffix}"
            sheet_name = (
                base_name[
                    : 31 - len(suffix_text)
                ]
                + suffix_text
            )
            suffix += 1

        used_sheet_names.add(sheet_name)

        worksheet = workbook.create_sheet(
            sheet_name
        )

        write_sheet(
            worksheet,
            group_data,
        )

        row_counts[sheet_name] = len(group_data)

    if len(workbook.sheetnames) != 11:
        raise AssertionError(
            f"Expected 11 sheets, generated "
            f"{len(workbook.sheetnames)}."
        )

    workbook.save(output_path)

    validation_workbook = load_workbook(
        output_path,
        read_only=True,
        data_only=False,
    )

    try:
        if len(validation_workbook.sheetnames) != 11:
            raise AssertionError(
                "Saved workbook does not contain "
                "exactly 11 sheets."
            )
    finally:
        validation_workbook.close()

    print(f"Created: {output_path}")
    print(f"Sheet count: {len(row_counts)}")

    for sheet_name, row_count in row_counts.items():
        print(
            f"- {sheet_name}: "
            f"{row_count} companies"
        )

    return output_path


def main() -> None:
    create_peer_comparison_workbook()


if __name__ == "__main__":
    main()
