"""Sprint 4 valuation engine: FCF yield and P/E valuation flags."""

from __future__ import annotations

from pathlib import Path
import sqlite3

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import get_settings

PROJECT_ROOT = Path(__file__).resolve().parents[2]
SETTINGS = get_settings()
DATABASE_PATH = Path(SETTINGS.database_path)
DEFAULT_OUTPUT_DIR = Path(SETTINGS.output_dir)


def _find_market_cap_excel() -> Path | None:
    candidates = [
        PROJECT_ROOT / "data" / "market_cap.xlsx",
        PROJECT_ROOT / "data" / "raw" / "market_cap.xlsx",
        PROJECT_ROOT / "data" / "raw" / "supporting datasets" / "market_cap.xlsx",
        PROJECT_ROOT / "supporting datasets" / "market_cap.xlsx",
        PROJECT_ROOT / "market_cap.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    for candidate in PROJECT_ROOT.rglob("market_cap.xlsx"):
        if ".venv" not in candidate.parts and ".git" not in candidate.parts:
            return candidate
    return None


def _load_market_cap(connection: sqlite3.Connection) -> pd.DataFrame:
    excel = _find_market_cap_excel()
    if excel is not None:
        data = pd.read_excel(excel)
        print(f"Market-cap source: {excel}")
        return data
    print(f"Market-cap source: SQLite table in {DATABASE_PATH}")
    return pd.read_sql_query("SELECT * FROM market_cap", connection)


def _style_excel(path: Path) -> None:
    workbook = load_workbook(path)
    sheet = workbook.active
    header_fill = PatternFill("solid", fgColor="1F4E78")
    caution_fill = PatternFill("solid", fgColor="F4CCCC")
    discount_fill = PatternFill("solid", fgColor="D9EAD3")
    fair_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    headers = {cell.value: cell.column for cell in sheet[1]}
    flag_col = headers.get("flag")
    if flag_col:
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=flag_col)
            if cell.value == "Caution":
                cell.fill = caution_fill
            elif cell.value == "Discount":
                cell.fill = discount_fill
            else:
                cell.fill = fair_fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in range(1, sheet.max_column + 1):
        max_length = max(len(str(sheet.cell(row=row, column=column).value or "")) for row in range(1, sheet.max_row + 1))
        sheet.column_dimensions[get_column_letter(column)].width = min(max(max_length + 2, 12), 34)
    workbook.save(path)


def build_valuation(output_dir: Path | str | None = None) -> tuple[pd.DataFrame, pd.DataFrame]:
    output_path = Path(output_dir) if output_dir is not None else DEFAULT_OUTPUT_DIR
    output_path.mkdir(parents=True, exist_ok=True)
    if not DATABASE_PATH.exists():
        raise FileNotFoundError(f"Database not found: {DATABASE_PATH}")

    with sqlite3.connect(DATABASE_PATH) as connection:
        companies = pd.read_sql_query(
            """
            SELECT c.company_id AS company_id, c.company_name, c.broad_sector AS sector
            FROM companies c
            ORDER BY c.company_id
            """,
            connection,
        )
        ratios = pd.read_sql_query(
            """
            SELECT company_id, CAST(year AS INTEGER) AS year, free_cash_flow_cr
            FROM financial_ratios
            """,
            connection,
        )
        market = _load_market_cap(connection)

    required_market_columns = {
        "company_id", "year", "market_cap_crore", "pe_ratio", "pb_ratio", "ev_ebitda"
    }
    missing = required_market_columns - set(market.columns)
    if missing:
        raise ValueError(f"market_cap data are missing columns: {sorted(missing)}")

    market = market.copy()
    market["company_id"] = market["company_id"].astype(str).str.strip().str.upper()
    market["year"] = pd.to_numeric(market["year"], errors="coerce")
    market = market.dropna(subset=["company_id", "year"])
    market["year"] = market["year"].astype(int)
    for column in ["market_cap_crore", "pe_ratio", "pb_ratio", "ev_ebitda"]:
        market[column] = pd.to_numeric(market[column], errors="coerce")

    ratios = ratios.copy()
    ratios["company_id"] = ratios["company_id"].astype(str).str.strip().str.upper()
    ratios["year"] = pd.to_numeric(ratios["year"], errors="coerce")
    ratios["free_cash_flow_cr"] = pd.to_numeric(ratios["free_cash_flow_cr"], errors="coerce")
    ratios = ratios.dropna(subset=["company_id", "year"])
    ratios["year"] = ratios["year"].astype(int)
    ratios = ratios.groupby(["company_id", "year"], as_index=False)["free_cash_flow_cr"].mean()

    latest_market = (
        market.sort_values(["company_id", "year"])
        .drop_duplicates("company_id", keep="last")
    )
    five_year_median = (
        market.sort_values(["company_id", "year"])
        .groupby("company_id", group_keys=False)
        .tail(5)
        .groupby("company_id")["pe_ratio"]
        .median()
        .rename("5yr_median_PE")
        .reset_index()
    )

    exact_fcf = latest_market[["company_id", "year"]].merge(
        ratios,
        on=["company_id", "year"],
        how="left",
    )
    latest_fcf = (
        ratios.sort_values(["company_id", "year"])
        .drop_duplicates("company_id", keep="last")
        [["company_id", "free_cash_flow_cr"]]
        .rename(columns={"free_cash_flow_cr": "fallback_fcf"})
    )
    exact_fcf = exact_fcf.merge(latest_fcf, on="company_id", how="left")
    exact_fcf["free_cash_flow_cr"] = exact_fcf["free_cash_flow_cr"].fillna(exact_fcf["fallback_fcf"])

    valuation = companies.merge(latest_market, on="company_id", how="left")
    valuation = valuation.merge(
        exact_fcf[["company_id", "free_cash_flow_cr"]],
        on="company_id",
        how="left",
    )
    valuation = valuation.merge(five_year_median, on="company_id", how="left")

    valuation["FCF_yield_pct"] = np.where(
        valuation["market_cap_crore"].gt(0),
        valuation["free_cash_flow_cr"] / valuation["market_cap_crore"] * 100,
        np.nan,
    )
    valuation["sector_median_pe"] = valuation.groupby("sector")["pe_ratio"].transform("median")
    valuation["PE_vs_sector_median_pct"] = np.where(
        valuation["sector_median_pe"].gt(0),
        (valuation["pe_ratio"] / valuation["sector_median_pe"] - 1) * 100,
        np.nan,
    )

    def assign_flag(row: pd.Series) -> str:
        pe = row["pe_ratio"]
        median = row["sector_median_pe"]
        if pd.isna(pe) or pd.isna(median) or median <= 0:
            return "Fair"
        if pe > median * 1.5:
            return "Caution"
        if pe < median * 0.7:
            return "Discount"
        return "Fair"

    valuation["flag"] = valuation.apply(assign_flag, axis=1)
    summary = valuation[
        [
            "company_id", "company_name", "sector", "pe_ratio", "pb_ratio", "ev_ebitda",
            "FCF_yield_pct", "5yr_median_PE", "PE_vs_sector_median_pct", "flag",
        ]
    ].rename(
        columns={
            "pe_ratio": "P/E",
            "pb_ratio": "P/B",
            "ev_ebitda": "EV/EBITDA",
        }
    )
    numeric = ["P/E", "P/B", "EV/EBITDA", "FCF_yield_pct", "5yr_median_PE", "PE_vs_sector_median_pct"]
    summary[numeric] = summary[numeric].round(2)

    if len(summary) != 92 or summary["company_id"].nunique() != 92:
        raise AssertionError(
            f"Expected 92 unique companies, found rows={len(summary)}, unique={summary['company_id'].nunique()}"
        )
    if not set(summary["flag"].dropna()).issubset({"Caution", "Discount", "Fair"}):
        raise AssertionError("Unexpected valuation flag value")

    flags = summary[summary["flag"].isin(["Caution", "Discount"])].copy()
    summary_file = output_path / "valuation_summary.xlsx"
    flags_file = output_path / "valuation_flags.csv"
    summary.to_excel(summary_file, index=False)
    _style_excel(summary_file)
    flags.to_csv(flags_file, index=False, encoding="utf-8-sig")

    print("Valuation module completed")
    print(f"Summary: {summary_file} ({len(summary)} rows)")
    print(f"Flags:   {flags_file} ({len(flags)} rows)")
    print(summary["flag"].value_counts().to_string())
    return summary, flags


if __name__ == "__main__":
    build_valuation()
